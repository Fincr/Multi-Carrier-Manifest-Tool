"""
Core manifest population engine.
"""

import os
from datetime import datetime
from typing import List, Tuple, Dict, Callable, Optional
from dataclasses import dataclass, field
import pandas as pd
from openpyxl import load_workbook

from carriers import get_carrier, ShipmentRecord
from carriers.spring import SpringCarrier
from carriers.landmark import LandmarkCarrier
from carriers.deutschepost import DeutschePostCarrier


@dataclass
class ProcessingResult:
    """Result of processing a carrier's data."""
    carrier_name: str
    output_file: str
    records_processed: int
    records_failed: int
    errors: List[str]
    success: bool
    po_number: str = ""
    additional_files: List[str] = field(default_factory=list)  # For carriers that generate multiple files
    deutschepost_data: Optional[object] = None  # For Deutsche Post carrier data


class ManifestEngine:
    """Engine for populating carrier manifests from carrier sheets."""
    
    def __init__(self, template_dir: str, output_dir: str):
        self.template_dir = template_dir
        self.output_dir = output_dir
        self.log_callback: Optional[Callable[[str], None]] = None
    
    def set_log_callback(self, callback: Callable[[str], None]):
        """Set callback for logging messages."""
        self.log_callback = callback
    
    def log(self, message: str):
        """Log a message."""
        if self.log_callback:
            self.log_callback(message)
        else:
            print(message)
    
    def load_carrier_sheet(self, filepath: str) -> Tuple[pd.DataFrame, str, str]:
        """
        Load carrier sheet and extract metadata.
        
        Returns:
            (data_df, po_number, carrier_name)
        """
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Extract metadata
        carrier_name = str(ws['B3'].value or "").strip()
        po_raw = ws['B4'].value
        po_number = str(int(po_raw)) if isinstance(po_raw, float) else str(po_raw or "")
        
        wb.close()
        
        # Load data section
        df = pd.read_excel(filepath, header=7)
        df = df.dropna(subset=['Country'])
        
        # Standardise column names
        df.columns = [str(c).strip() for c in df.columns]
        
        # Ensure required columns exist
        required = ['Country', 'Service', 'Format', 'Items', 'Weight (KG)']
        for col in required:
            if col not in df.columns:
                raise ValueError(f"Missing required column: {col}")
        
        df = df[required].copy()
        df['Items'] = pd.to_numeric(df['Items'], errors='coerce').fillna(0).astype(int)
        df['Weight (KG)'] = pd.to_numeric(df['Weight (KG)'], errors='coerce').fillna(0.0)
        
        return df, po_number, carrier_name
    
    def group_by_carrier(self, df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """
        Group data by carrier.
        
        For now, assumes single carrier per sheet based on B3 value.
        Could be extended to support multi-carrier sheets.
        """
        # Currently returns single group - carrier determined at sheet level
        return {'default': df}
    
    def process_carrier(self, carrier_name: str, data: pd.DataFrame, 
                       po_number: str, max_errors: int = 5) -> ProcessingResult:
        """
        Process data for a single carrier.
        
        Args:
            carrier_name: Name of carrier (must match registry)
            data: DataFrame with shipment records
            po_number: PO number for manifest
            max_errors: Stop processing if errors exceed this count
        
        Returns:
            ProcessingResult with outcome details
        """
        errors = []
        records_processed = 0
        records_failed = 0
        
        try:
            carrier = get_carrier(carrier_name)
        except ValueError as e:
            return ProcessingResult(
                carrier_name=carrier_name,
                output_file="",
                records_processed=0,
                records_failed=len(data),
                errors=[str(e)],
                success=False,
                po_number=po_number
            )
        
        # Handle Deutsche Post carrier - special processing (no template)
        if isinstance(carrier, DeutschePostCarrier):
            # Deutsche Post is handled differently at sheet level
            # This shouldn't be called directly - use process_deutschepost_sheet instead
            return ProcessingResult(
                carrier_name="Deutsche Post",
                output_file="",
                records_processed=0,
                records_failed=0,
                errors=["Deutsche Post requires process_deutschepost_sheet method"],
                success=False,
                po_number=po_number
            )
            
        # Load template
        template_path = os.path.join(self.template_dir, carrier.template_filename)
        if not os.path.exists(template_path):
            return ProcessingResult(
                carrier_name=carrier_name,
                output_file="",
                records_processed=0,
                records_failed=len(data),
                errors=[f"Template not found: {template_path}"],
                success=False,
                po_number=po_number
            )
            
        self.log(f"Loading template: {carrier.template_filename}")
        
        # Handle Spring carrier differently - it generates order lines
        if isinstance(carrier, SpringCarrier):
            return self._process_spring_carrier(carrier, data, po_number, template_path, max_errors)
        
        # Handle Landmark carrier - generates CSV files
        if isinstance(carrier, LandmarkCarrier):
            return self._process_landmark_carrier(carrier, data, po_number, template_path, max_errors)
        
        # Standard carrier processing (Asendia, PostNord)
        wb = load_workbook(template_path)
        
        # Build country index
        country_index = carrier.build_country_index(wb)
        self.log(f"Built index with {len(country_index)} countries")
        
        # Set metadata
        shipment_date = datetime.now().strftime("%Y-%m-%d")
        carrier.set_metadata(wb, po_number, shipment_date)
        self.log(f"Set PO: {po_number}, Date: {shipment_date}")
        
        # Process each record
        for idx, row in data.iterrows():
            record = ShipmentRecord(
                country=str(row['Country']),
                service=str(row['Service']),
                format=str(row['Format']),
                items=int(row['Items']),
                weight=float(row['Weight (KG)'])
            )
            
            result = carrier.place_record(wb, record, country_index)
            
            if result.success:
                records_processed += 1
            else:
                records_failed += 1
                errors.append(result.error_message)
                self.log(f"  ⚠ {result.error_message}")
                
                if len(errors) > max_errors:
                    errors.append(f"Stopping: exceeded {max_errors} errors")
                    self.log(f"  ✗ Stopping: exceeded {max_errors} errors")
                    break
        
        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_carrier = str(carrier_name).replace(" ", "_").replace("/", "-")
        safe_po = str(po_number)
        output_filename = f"{safe_carrier}_{safe_po}_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, output_filename)
        
        # Save
        wb.save(output_path)
        wb.close()
        self.log(f"Saved: {output_filename}")
        
        return ProcessingResult(
            carrier_name=carrier_name,
            output_file=output_path,
            records_processed=records_processed,
            records_failed=records_failed,
            errors=errors,
            success=len(errors) <= max_errors,
            po_number=po_number
        )
    
    def _process_spring_carrier(self, carrier: SpringCarrier, data: pd.DataFrame,
                                 po_number: str, template_path: str, 
                                 max_errors: int = 5) -> ProcessingResult:
        """
        Process Spring carrier - generates order lines instead of populating cells.
        """
        errors = []
        records_processed = 0
        records_failed = 0
        
        # Clear any previous order lines
        carrier.clear_order_lines()
        
        # Set PO number
        carrier.set_metadata(None, po_number, "")
        self.log(f"Set PO: {po_number}")
        
        # Process each record into order lines
        for idx, row in data.iterrows():
            record = ShipmentRecord(
                country=str(row['Country']),
                service=str(row['Service']),
                format=str(row['Format']),
                items=int(row['Items']),
                weight=float(row['Weight (KG)'])
            )
            
            result = carrier.place_record(None, record, {})
            
            if result.success:
                records_processed += 1
            else:
                records_failed += 1
                errors.append(result.error_message)
                self.log(f"  ⚠ {result.error_message}")
                
                if len(errors) > max_errors:
                    errors.append(f"Stopping: exceeded {max_errors} errors")
                    self.log(f"  ✗ Stopping: exceeded {max_errors} errors")
                    break
        
        # Generate output filename - save to output folder for debugging
        # The upload file will be deleted after successful portal download/print
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_po = str(po_number)
        output_filename = f"Spring_Upload_{safe_po}_{timestamp}.xlsx"
        
        # Save to output directory (will be deleted after successful portal processing)
        output_path = os.path.join(self.output_dir, output_filename)
        
        # Write order lines to manifest
        order_lines = carrier.get_order_lines()
        priority_count = len([line for line in order_lines if line.product_code == '1MI'])
        economy_count = len([line for line in order_lines if line.product_code == '2MI'])
        self.log(f"Generated {len(order_lines)} order lines (Priority: {priority_count}, Economy: {economy_count})")
        carrier.write_manifest(template_path, output_path)
        self.log(f"Saved upload file: {output_filename}")
        
        return ProcessingResult(
            carrier_name="Spring",
            output_file=output_path,
            records_processed=records_processed,
            records_failed=records_failed,
            errors=errors,
            success=len(errors) <= max_errors,
            po_number=po_number
        )
    
    def _process_landmark_carrier(self, carrier: LandmarkCarrier, data: pd.DataFrame,
                                   po_number: str, iso_code_path: str,
                                   max_errors: int = 5) -> ProcessingResult:
        """
        Process Landmark Global carrier - generates CSV upload files.
        
        Args:
            carrier: LandmarkCarrier instance
            data: DataFrame with shipment records
            po_number: PO number for manifest
            iso_code_path: Path to UploadCodeList_-_Citipost.xls
            max_errors: Stop processing if errors exceed this count
            
        Returns:
            ProcessingResult with outcome details
        """
        errors = []
        records_processed = 0
        records_failed = 0
        
        # Clear any previous order lines
        carrier.clear_order_lines()
        
        # Load ISO codes
        if not carrier.load_iso_codes(iso_code_path):
            return ProcessingResult(
                carrier_name="Landmark Global",
                output_file="",
                records_processed=0,
                records_failed=len(data),
                errors=carrier.errors,
                success=False,
                po_number=po_number
            )
        
        self.log(f"Loaded ISO code mappings from: {os.path.basename(iso_code_path)}")
        
        # Set PO number and date
        shipment_date = datetime.now().strftime("%Y-%m-%d")
        carrier.set_metadata(None, po_number, shipment_date)
        self.log(f"Set PO: {po_number}, Date: {shipment_date}")
        
        # Process each record into order lines
        for idx, row in data.iterrows():
            record = ShipmentRecord(
                country=str(row['Country']),
                service=str(row['Service']),
                format=str(row['Format']),
                items=int(row['Items']),
                weight=float(row['Weight (KG)'])
            )
            
            result = carrier.place_record(None, record, {})
            
            if result.success:
                records_processed += 1
            else:
                records_failed += 1
                errors.append(result.error_message)
                self.log(f"  ⚠ {result.error_message}")
                
                if len(errors) > max_errors:
                    errors.append(f"Stopping: exceeded {max_errors} errors")
                    self.log(f"  ✗ Stopping: exceeded {max_errors} errors")
                    break
        
        # Get summary
        summary = carrier.get_summary()
        for service_type, stats in summary.items():
            self.log(f"  {service_type}: {stats['rows']} rows, {stats['total_pieces']} pieces, {stats['total_weight_kg']} kg")
        
        # Write upload files to output directory
        files_created = carrier.write_upload_files(self.output_dir)
        
        self.log(f"Files created: {len(files_created)}")
        
        if files_created:
            for filepath in files_created:
                self.log(f"Saved: {os.path.basename(filepath)}")
            
            # Primary file is the first one, additional files stored separately
            primary_file = files_created[0]
            additional_files = files_created[1:] if len(files_created) > 1 else []
            
            self.log(f"Primary file: {os.path.basename(primary_file)}")
            if additional_files:
                self.log(f"Additional files: {[os.path.basename(f) for f in additional_files]}")
            else:
                self.log("No additional files (single service type only)")
            
            return ProcessingResult(
                carrier_name="Landmark Global",
                output_file=primary_file,
                records_processed=records_processed,
                records_failed=records_failed,
                errors=errors,
                success=len(errors) <= max_errors,
                po_number=po_number,
                additional_files=additional_files
            )
        else:
            errors.append("No upload files generated")
            return ProcessingResult(
                carrier_name="Landmark Global",
                output_file="",
                records_processed=records_processed,
                records_failed=records_failed,
                errors=errors,
                success=False,
                po_number=po_number
            )
    
    def _process_deutschepost_carrier(self, carrier: DeutschePostCarrier, 
                                        carrier_sheet_path: str) -> ProcessingResult:
        """
        Process Deutsche Post carrier - modifies and saves carrier sheet.
        
        Unlike other carriers, Deutsche Post:
        1. Doesn't use a template - uses the carrier sheet directly
        2. Removes the (EMB) Manifest sheet
        3. Saves modified carrier sheet to output directory
        4. Portal automation fills a form instead of uploading CSV
        
        Args:
            carrier: DeutschePostCarrier instance
            carrier_sheet_path: Path to original carrier sheet
            
        Returns:
            ProcessingResult with outcome details
        """
        errors = []
        
        try:
            # Process carrier sheet (extract data, remove EMB sheet, save)
            output_path, extracted_data = carrier.process_carrier_sheet(
                carrier_sheet_path, 
                self.output_dir
            )
            
            self.log("Extracted data:")
            self.log(f"  PO Number: {extracted_data.po_number}")
            self.log(f"  Total Weight: {extracted_data.total_weight} kg")
            self.log(f"  Total Items: {extracted_data.total_items}")
            self.log(f"  Formats: {', '.join(extracted_data.formats)}")
            self.log(f"  Item Format for portal: {carrier.get_item_format(extracted_data.formats)}")
            self.log("")
            self.log("Removed (EMB) Manifest sheet")
            self.log(f"Saved: {os.path.basename(output_path)}")
            
            return ProcessingResult(
                carrier_name="Deutsche Post",
                output_file=output_path,
                records_processed=extracted_data.total_items,
                records_failed=0,
                errors=errors,
                success=True,
                po_number=extracted_data.po_number,
                deutschepost_data=extracted_data
            )
            
        except Exception as e:
            errors.append(str(e))
            return ProcessingResult(
                carrier_name="Deutsche Post",
                output_file="",
                records_processed=0,
                records_failed=0,
                errors=errors,
                success=False,
                po_number=""
            )
    
    def process_sheet(self, carrier_sheet_path: str, 
                      max_errors: int = 5) -> List[ProcessingResult]:
        """
        Process a carrier sheet, populating manifests for all carriers found.
        
        Args:
            carrier_sheet_path: Path to internal carrier sheet
            max_errors: Stop processing each carrier if errors exceed this
        
        Returns:
            List of ProcessingResult, one per carrier
        """
        results = []
        
        self.log(f"Loading carrier sheet: {os.path.basename(carrier_sheet_path)}")
        
        try:
            data, po_number, carrier_name = self.load_carrier_sheet(carrier_sheet_path)
        except Exception as e:
            return [ProcessingResult(
                carrier_name="Unknown",
                output_file="",
                records_processed=0,
                records_failed=0,
                errors=[f"Failed to load carrier sheet: {str(e)}"],
                success=False,
                po_number=""
            )]
        
        self.log(f"Carrier: {carrier_name}")
        self.log(f"PO: {po_number}")
        self.log(f"Records: {len(data)}")
        
        # Check for zero records - cancel processing if carrier sheet is empty
        if len(data) == 0:
            self.log("")
            self.log("⚠ ZERO RECORDS DETECTED")
            self.log("The carrier sheet contains no data rows.")
            self.log("Processing cancelled.")
            return [ProcessingResult(
                carrier_name=carrier_name,
                output_file="",
                records_processed=0,
                records_failed=0,
                errors=["Carrier sheet contains zero records - processing cancelled"],
                success=False,
                po_number=po_number
            )]
        
        # Check for Deutsche Post - requires special handling
        carrier_lower = carrier_name.lower().strip()
        if 'deutsche' in carrier_lower:
            carrier = DeutschePostCarrier()
            result = self._process_deutschepost_carrier(carrier, carrier_sheet_path)
            results.append(result)
            return results
        
        # Process the carrier (standard flow)
        result = self.process_carrier(carrier_name, data, po_number, max_errors)
        results.append(result)
        
        return results
