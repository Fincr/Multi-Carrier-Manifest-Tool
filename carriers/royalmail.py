"""
Royal Mail International 2026 carrier handler.

A single carrier sheet may contain both Flats and Letters rows for Ireland.
Data is extracted, bucketed by format, and submitted to the Royal Mail OBA
portal which generates the manifest. There is no manifest template.
"""

import os
from datetime import datetime
from typing import Dict, Tuple, Optional
from dataclasses import dataclass

from openpyxl import load_workbook

from .base import BaseCarrier, ShipmentRecord, PlacementResult


@dataclass
class RoyalMailData:
    """Data extracted from a Royal Mail International carrier sheet.

    Weight fields store total weight in KG from the carrier sheet.
    The portal requires average item weight in grams, computed via
    avg_weight_grams().
    """
    po_number: str
    flats_items: int = 0
    flats_weight: float = 0.0      # total weight in KG
    letters_items: int = 0
    letters_weight: float = 0.0    # total weight in KG

    def avg_weight_grams(self, format_type: str) -> int:
        """Average item weight in grams for a format (Letters or Flats).

        The OBA portal requires this as a whole number in grams.
        """
        if format_type == 'letters' and self.letters_items > 0:
            return round(self.letters_weight * 1000 / self.letters_items)
        elif format_type == 'flats' and self.flats_items > 0:
            return round(self.flats_weight * 1000 / self.flats_items)
        return 0


class RoyalMailCarrier(BaseCarrier):
    """
    Handler for Royal Mail International 2026.

    Like Deutsche Post, this carrier has no manifest template.
    Data is extracted from the carrier sheet and submitted to the
    Royal Mail OBA portal, which generates the manifest.
    A single sheet may contain both Flats and Letters rows.
    """

    carrier_name = "Royal Mail International 2026"
    template_filename = ""  # No template — portal generates the manifest

    # Ireland country name variations
    IRELAND_NAMES = {
        'ireland', 'republic of ireland', 'eire', 'ie',
        'ireland, republic of', 'roi',
    }

    def __init__(self):
        super().__init__()
        self.country_mapping = {
            'Republic of Ireland': 'Ireland',
            'Eire': 'Ireland',
            'IE': 'Ireland',
            'Ireland, Republic of': 'Ireland',
            'ROI': 'Ireland',
        }

    def build_country_index(self, workbook) -> Dict[str, dict]:
        """Not used — Royal Mail has no template."""
        return {}

    def get_cell_positions(self, country_info: dict, format_type: str) -> Tuple[int, int]:
        """Not used — Royal Mail has no template."""
        raise NotImplementedError("Royal Mail does not use cell-based manifests")

    def set_metadata(self, workbook, po_number: str, shipment_date: str) -> None:
        """Not used — Royal Mail has no template."""
        pass

    def place_record(self, workbook, record: ShipmentRecord, country_index: dict) -> PlacementResult:
        """Not used — Royal Mail processes data via extract_data instead."""
        raise NotImplementedError("Royal Mail does not use place_record — use process_carrier_sheet")

    def process_carrier_sheet(
        self,
        carrier_sheet_path: str,
        output_dir: str,
        log_callback=None
    ) -> Tuple[str, RoyalMailData]:
        """
        Extract data from a Royal Mail carrier sheet and save to output dir.

        Args:
            carrier_sheet_path: Path to the original carrier sheet
            output_dir: Directory to save the processed sheet
            log_callback: Optional logging function

        Returns:
            (output_path, extracted_data)
        """
        def log(msg):
            if log_callback:
                log_callback(msg)

        wb = load_workbook(carrier_sheet_path, data_only=True)
        ws = wb.active

        # Extract metadata
        carrier_name = str(ws['B3'].value or "").strip()
        po_raw = ws['B4'].value
        po_number = str(int(po_raw)) if isinstance(po_raw, float) else str(po_raw or "")

        # Read data rows (header at row 8, data from row 9)
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=8, column=col).value
            if val:
                headers[str(val).strip()] = col

        flats_items = 0
        flats_weight = 0.0
        letters_items = 0
        letters_weight = 0.0
        non_ireland_countries = []
        unknown_formats = []

        row = 9
        while row <= ws.max_row:
            country_val = ws.cell(row=row, column=headers.get('Country', 1)).value
            if country_val is None or str(country_val).strip() == '':
                row += 1
                continue

            country = str(country_val).strip()
            format_val = str(ws.cell(row=row, column=headers.get('Format', 4)).value or "").strip()
            items_val = ws.cell(row=row, column=headers.get('Items', 5)).value
            weight_val = ws.cell(row=row, column=headers.get('Weight (KG)', 6)).value

            # Validate country is Ireland
            mapped_country = self.map_country(country)
            if mapped_country.lower() not in self.IRELAND_NAMES:
                non_ireland_countries.append(country)

            # Parse numeric values
            try:
                items = int(items_val) if items_val not in (None, '', ' ') else 0
            except (ValueError, TypeError):
                items = 0

            try:
                weight = float(weight_val) if weight_val not in (None, '', ' ') else 0.0
            except (ValueError, TypeError):
                weight = 0.0

            # Bucket by format
            normalised_format = self.normalise_format(format_val)
            if normalised_format == 'Flats':
                flats_items += items
                flats_weight += weight
            elif normalised_format == 'Letters':
                letters_items += items
                letters_weight += weight
            else:
                unknown_formats.append(format_val)

            row += 1

        # Log warnings
        if non_ireland_countries:
            unique = set(non_ireland_countries)
            log(f"  ⚠ Non-Ireland countries found: {', '.join(unique)}")

        if unknown_formats:
            unique = set(unknown_formats)
            log(f"  ⚠ Unrecognised formats: {', '.join(unique)}")

        # Build result data
        data = RoyalMailData(
            po_number=po_number,
            flats_items=flats_items,
            flats_weight=round(flats_weight, 3),
            letters_items=letters_items,
            letters_weight=round(letters_weight, 3),
        )

        # Save carrier sheet to output directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = carrier_name.replace(" ", "_").replace("/", "-")
        output_filename = f"{safe_name}_{po_number}_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        wb.save(output_path)
        wb.close()

        return output_path, data


