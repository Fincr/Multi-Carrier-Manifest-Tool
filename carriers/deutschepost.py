"""
Deutsche Post carrier handler.

Deutsche Post uses a simpler workflow:
1. Remove (EMB) Manifest sheet from carrier sheet
2. Save modified carrier sheet to output location
3. Print to default printer
4. Open portal and register manifest via form (no CSV upload)

Portal workflow:
- Login to portal
- Navigate: Ship > Prepare Airway Bills > Print Airway Bill
- Fill form: Contact name, Job reference (PO), Item format, Total weight
- Submit and download manifest PDF
"""

from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
from datetime import datetime
import os
from openpyxl import load_workbook

from .base import BaseCarrier, ShipmentRecord, PlacementResult


@dataclass
class DeutschePostData:
    """Data extracted from Deutsche Post carrier sheet."""
    po_number: str
    total_weight: float
    total_items: int
    formats: set  # Set of format types (Letters, Flats, Packets)
    

class DeutschePostCarrier(BaseCarrier):
    """Handler for Deutsche Post manifests."""
    
    carrier_name = "Deutsche Post"
    template_filename = ""  # No template - uses carrier sheet directly
    
    # Portal configuration
    PORTAL_URL = "https://packet.deutschepost.com/webapp/index.xhtml"
    # Note: Credentials loaded from environment/.env via core.credentials module
    
    # Format mapping for portal dropdown
    # Portal options are: P, G, E, mixed (P/G/E)
    # P = Letters, G = Flats/Large Letters, E = Packets
    FORMAT_MAP = {
        frozenset(['Letters']): 'P',
        frozenset(['Flats']): 'G',
        frozenset(['Packets']): 'E',
        # Any combination with multiple formats = Mixed
    }
    
    def __init__(self):
        super().__init__()
        self._carrier_sheet_path = ""
        self._extracted_data: Optional[DeutschePostData] = None
    
    def extract_data(self, carrier_sheet_path: str) -> DeutschePostData:
        """
        Extract required data from Deutsche Post carrier sheet.
        
        Args:
            carrier_sheet_path: Path to the carrier sheet
            
        Returns:
            DeutschePostData with PO, weight, items, formats
        """
        wb = load_workbook(carrier_sheet_path, data_only=True)
        ws = wb['Manifest']
        
        # Get PO number from B4
        po_raw = ws.cell(row=4, column=2).value
        po_number = str(int(po_raw)) if isinstance(po_raw, (int, float)) else str(po_raw or "")
        
        # Find totals row (last row with data in Items/Weight columns)
        total_weight = 0.0
        total_items = 0
        formats = set()
        
        # Data starts at row 9, iterate until we hit empty country
        for row in range(9, ws.max_row + 1):
            country = ws.cell(row=row, column=1).value
            
            # Check if this is a data row or totals row
            if country is None or str(country).strip() == '':
                # This should be the totals row
                items_val = ws.cell(row=row, column=5).value
                weight_val = ws.cell(row=row, column=6).value
                if items_val is not None:
                    total_items = int(items_val)
                if weight_val is not None:
                    total_weight = float(weight_val)
                break
            else:
                # This is a data row - collect format
                fmt = ws.cell(row=row, column=4).value
                if fmt:
                    formats.add(self.normalise_format(str(fmt)))
        
        wb.close()
        
        return DeutschePostData(
            po_number=po_number,
            total_weight=round(total_weight, 3),
            total_items=total_items,
            formats=formats
        )
    
    def get_item_format(self, formats: set) -> str:
        """
        Determine the item format for the portal dropdown.
        
        Args:
            formats: Set of format types present
            
        Returns:
            Format string for portal
        """
        formats_frozen = frozenset(formats)
        
        # Check for single format
        if formats_frozen in self.FORMAT_MAP:
            return self.FORMAT_MAP[formats_frozen]
        
        # Multiple formats = Mixed
        return "mixed (P/G/E)"
    
    def process_carrier_sheet(self, input_path: str, output_dir: str) -> Tuple[str, DeutschePostData]:
        """
        Process Deutsche Post carrier sheet:
        1. Extract data
        2. Remove (EMB) Manifest sheet
        3. Save to output directory
        
        Args:
            input_path: Path to original carrier sheet
            output_dir: Directory to save processed file
            
        Returns:
            Tuple of (output_filepath, extracted_data)
        """
        # Extract data first
        data = self.extract_data(input_path)
        self._extracted_data = data
        
        # Load workbook for modification
        wb = load_workbook(input_path)
        
        # Remove (EMB) Manifest sheet if it exists
        if '(EMB) Manifest' in wb.sheetnames:
            del wb['(EMB) Manifest']
        
        # Generate output filename with standard naming convention
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        input_filename = os.path.basename(input_path)
        # Remove extension and add timestamp
        base_name = os.path.splitext(input_filename)[0]
        output_filename = f"{base_name}_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        # Save
        wb.save(output_path)
        wb.close()
        
        return output_path, data
    
    # Required abstract method implementations (not used for Deutsche Post)
    def build_country_index(self, workbook) -> Dict[str, dict]:
        """Not used for Deutsche Post - no manifest template."""
        return {}
    
    def get_cell_positions(self, country_info: dict, format_type: str) -> Tuple[int, int]:
        """Not used for Deutsche Post - no cell placement."""
        raise NotImplementedError("Deutsche Post does not use cell placement")
    
    def set_metadata(self, workbook, po_number: str, shipment_date: str) -> None:
        """Not used for Deutsche Post - no manifest template."""
        pass
