"""
Spring Global Delivery Solutions manifest handler.

Spring uses a different manifest format - a flat upload sheet where each
country/format combination is a separate row, rather than a matrix layout.
"""

from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
from openpyxl import load_workbook
from .base import BaseCarrier, ShipmentRecord, PlacementResult


@dataclass
class SpringOrderLine:
    """A single order line for Spring manifest."""
    customer_number: str
    customer_ref_1: str  # PO number
    customer_ref_2: str
    quote_ref: str
    count_sort: str  # Y/N
    pre_franked: str  # Y/N
    product_code: str  # 1MI (Priority) or 2MI (Economy)
    nr_satchels: Optional[int]
    nr_bags: Optional[int]
    nr_boxes: Optional[int]
    nr_pallets: Optional[int]
    nr_trays: Optional[int]
    destination_code: str  # ISO code or regional code
    format_code: str  # L, P, B, G, N, E
    weightbreak_from: Optional[float]
    weightbreak_to: Optional[float]
    nr_items: int
    weight_kg: float


class SpringCarrier(BaseCarrier):
    """Handler for Spring Global Delivery Solutions manifests."""
    
    carrier_name = "Spring"
    template_filename = "MailOrderTemplate.xlsx"
    
    # Static configuration
    CUSTOMER_NUMBER = "100007596"
    COUNT_SORT = "N"
    PRE_FRANKED = "Y"
    
    # Service to product code mapping
    SERVICE_TO_PRODUCT = {
        'Priority': '1MI',
        'Economy': '2MI',
    }
    
    # EU countries (use B/L/N format codes)
    EU_COUNTRIES = {
        'Austria', 'Belgium', 'Bulgaria', 'Croatia', 'Cyprus', 'Czech Republic',
        'Denmark', 'Estonia', 'Finland', 'France', 'Germany', 'Greece', 'Hungary',
        'Iceland', 'Ireland', 'Italy', 'Latvia', 'Lithuania', 'Luxembourg', 'Malta',
        'Netherlands', 'Norway', 'Poland', 'Portugal', 'Romania', 'Serbia',
        'Slovakia', 'Slovenia', 'Spain', 'Sweden', 'Switzerland',
        # Regional EU codes
        'Rest of Europe non EU',
    }
    
    # Format code mapping: (is_EU) -> carrier_format -> spring_code
    FORMAT_CODES = {
        True: {   # EU
            'Letters': 'L',
            'Flats': 'B',     # BOXABLE
            'Packets': 'N',   # NON BOXABLE
        },
        False: {  # ROW
            'Letters': 'P',
            'Flats': 'G',
            'Packets': 'E',
        }
    }
    
    # Direct country name to destination code mapping
    COUNTRY_TO_CODE = {
        'Africa': 'AFR',
        'Argentina': 'AR',
        'Australia': 'AU',
        'Austria': 'AT',
        'Belarus': 'BY',
        'Belgium': 'BE',
        'Brazil': 'BR',
        'Bulgaria': 'BG',
        'Canada': 'CA',
        'Central and South America': 'CSA',
        'Chile': 'CL',
        'China': 'CN',
        'Croatia': 'HR',
        'Cyprus': 'CY',
        'Czech Republic': 'CZ',
        'Denmark': 'DK',
        'Estonia': 'EE',
        'Far East': 'FEA',
        'Finland': 'FI',
        'France': 'FR',
        'Germany': 'DE',
        'Greece': 'GR',
        'Hong Kong': 'HK',
        'Hungary': 'HU',
        'Iceland': 'IS',
        'India': 'IN',
        'Indonesia': 'ID',
        'Ireland': 'IE',
        'Israel': 'IL',
        'Italy': 'IT',
        'Japan': 'JP',
        'Latvia': 'LV',
        'Lithuania': 'LT',
        'Luxembourg': 'LU',
        'Malaysia': 'MY',
        'Malta': 'MT',
        'Mexico': 'MX',
        'Middle East': 'MEA',
        'Netherlands': 'NL',
        'New Zealand': 'NZ',
        'Norway': 'NO',
        'Poland': 'PL',
        'Portugal': 'PT',
        'Rest of Europe non EU': 'EUR',
        'Rest of World': 'ROW',
        'Romania': 'RO',
        'Russian Federation': 'RU',
        'Saudi Arabia': 'SA',
        'Serbia': 'RS',
        'Singapore': 'SG',
        'Slovakia': 'SK',
        'Slovenia': 'SI',
        'South Africa': 'ZA',
        'South Korea': 'KR',
        'Spain': 'ES',
        'Sweden': 'SE',
        'Switzerland': 'CH',
        'Taiwan': 'TW',
        'Thailand': 'TH',
        'Turkey': 'TR',
        'Ukraine': 'UA',
        'United Arab Emirates': 'AE',
        'United States of America': 'US',
        'United States': 'US',
    }
    
    # Regional fallback mapping for countries not directly supported
    REGIONAL_FALLBACKS = {
        # Europe non-EU
        'Aland Islands': 'EUR',
        'Gibraltar': 'EUR',
        'Monaco': 'EUR',
        'Liechtenstein': 'EUR',
        'Andorra': 'EUR',
        'San Marino': 'EUR',
        'Vatican City': 'EUR',
        'Faroe Islands': 'EUR',
        'Greenland': 'EUR',
        'Moldova': 'EUR',
        'North Macedonia': 'EUR',
        'Montenegro': 'EUR',
        'Albania': 'EUR',
        'Bosnia and Herzegovina': 'EUR',
        'Kosovo': 'EUR',
        
        # Middle East
        'Qatar': 'MEA',
        'Kuwait': 'MEA',
        'Bahrain': 'MEA',
        'Oman': 'MEA',
        'Jordan': 'MEA',
        'Lebanon': 'MEA',
        'Iraq': 'MEA',
        'Iran': 'MEA',
        'Yemen': 'MEA',
        'Syria': 'MEA',
        
        # Africa
        'Ghana': 'AFR',
        'Eswatini': 'AFR',
        'Libya': 'AFR',
        'Egypt': 'AFR',
        'Morocco': 'AFR',
        'Tunisia': 'AFR',
        'Algeria': 'AFR',
        'Nigeria': 'AFR',
        'Kenya': 'AFR',
        'Tanzania': 'AFR',
        'Uganda': 'AFR',
        'Ethiopia': 'AFR',
        'Senegal': 'AFR',
        'Ivory Coast': 'AFR',
        "Côte d'Ivoire": 'AFR',
        'Cameroon': 'AFR',
        'Zimbabwe': 'AFR',
        'Zambia': 'AFR',
        'Botswana': 'AFR',
        'Namibia': 'AFR',
        'Mozambique': 'AFR',
        'Madagascar': 'AFR',
        'Mauritius': 'AFR',
        'Reunion': 'AFR',
        'Seychelles': 'AFR',
        'Somalia': 'AFR',
        
        # Central and South America
        'Guatemala': 'CSA',
        'Guyana': 'CSA',
        'Honduras': 'CSA',
        'Nicaragua': 'CSA',
        'El Salvador': 'CSA',
        'Costa Rica': 'CSA',
        'Panama': 'CSA',
        'Colombia': 'CSA',
        'Ecuador': 'CSA',
        'Peru': 'CSA',
        'Bolivia': 'CSA',
        'Paraguay': 'CSA',
        'Uruguay': 'CSA',
        'Venezuela': 'CSA',
        'Suriname': 'CSA',
        'French Guiana': 'CSA',
        
        # Caribbean -> ROW (no specific Caribbean region)
        'Anguilla': 'ROW',
        'British Virgin Islands': 'ROW',
        'Canary Islands': 'ES',  # Part of Spain
        'Cayman Islands': 'ROW',
        'Saint Barthelemy': 'ROW',
        'Saint Helena': 'ROW',
        'Saint Kitts and Nevis': 'ROW',
        'Saint Vincent and the Grenadines': 'ROW',
        'Turks and Caicos Islands': 'ROW',
        'Bermuda': 'ROW',
        'Bahamas': 'ROW',
        'Jamaica': 'ROW',
        'Barbados': 'ROW',
        'Trinidad and Tobago': 'ROW',
        'Antigua and Barbuda': 'ROW',
        'Dominica': 'ROW',
        'Grenada': 'ROW',
        'Saint Lucia': 'ROW',
        'Martinique': 'ROW',
        'Guadeloupe': 'ROW',
        'Aruba': 'ROW',
        'Curacao': 'ROW',
        'Puerto Rico': 'US',  # US territory
        
        # Far East
        'Cook Islands': 'FEA',
        'Vietnam': 'FEA',
        'Philippines': 'FEA',
        'Cambodia': 'FEA',
        'Myanmar': 'FEA',
        'Laos': 'FEA',
        'Brunei': 'FEA',
        'Fiji': 'FEA',
        'Papua New Guinea': 'FEA',
        'Samoa': 'FEA',
        'Tonga': 'FEA',
        'Vanuatu': 'FEA',
        
        # Other ROW
        'Falkland Islands': 'ROW',
        'Pakistan': 'ROW',
        'Bangladesh': 'ROW',
        'Sri Lanka': 'ROW',
        'Nepal': 'ROW',
        'Afghanistan': 'ROW',
        'Maldives': 'ROW',
    }
    
    def __init__(self):
        super().__init__()
        self._order_lines: List[SpringOrderLine] = []
        self._po_number = ""
    
    def get_destination_code(self, country: str) -> Optional[str]:
        """Get Spring destination code for a country."""
        # Try direct mapping first
        if country in self.COUNTRY_TO_CODE:
            return self.COUNTRY_TO_CODE[country]
        
        # Try regional fallback
        if country in self.REGIONAL_FALLBACKS:
            return self.REGIONAL_FALLBACKS[country]
        
        return None
    
    def is_eu_destination(self, destination_code: str) -> bool:
        """Check if destination code is EU (uses B/L/N format codes).
        
        Note: EUR (Rest of Europe non EU) uses ROW format codes (P/G/E), not EU codes.
        """
        eu_codes = {'AT', 'BE', 'BG', 'HR', 'CY', 'CZ', 'DK', 'EE', 'FI', 'FR',
                    'DE', 'GR', 'HU', 'IS', 'IE', 'IT', 'LV', 'LT', 'LU', 'MT',
                    'NL', 'NO', 'PL', 'PT', 'RO', 'RS', 'SK', 'SI', 'ES', 'SE',
                    'CH'}  # EUR excluded - uses ROW format codes
        return destination_code in eu_codes
    
    def get_format_code(self, format_type: str, destination_code: str) -> str:
        """Get Spring format code based on format type and destination."""
        is_eu = self.is_eu_destination(destination_code)
        format_map = self.FORMAT_CODES[is_eu]
        return format_map.get(format_type, 'P')  # Default to P if unknown
    
    def build_country_index(self, workbook) -> Dict[str, dict]:
        """
        Spring doesn't use a pre-built country index in the same way.
        We build order lines dynamically. Return empty dict - not used.
        """
        return {}
    
    def get_cell_positions(self, country_info: dict, format_type: str) -> Tuple[int, int]:
        """Not used for Spring - we build order lines instead."""
        raise NotImplementedError("Spring uses order line generation, not cell placement")
    
    def set_metadata(self, workbook, po_number: str, shipment_date: str) -> None:
        """Store PO number for order line generation."""
        self._po_number = po_number
    
    def place_record(self, workbook, record: ShipmentRecord, country_index: dict) -> PlacementResult:
        """
        Convert a shipment record to a Spring order line.
        Instead of placing in cells, we accumulate order lines.
        """
        # Get destination code
        dest_code = self.get_destination_code(record.country)
        if not dest_code:
            return PlacementResult(
                success=False,
                error_message=f"No destination code mapping for: {record.country}"
            )
        
        # Get service/product code
        service = self.normalise_service(record.service)
        product_code = self.SERVICE_TO_PRODUCT.get(service)
        if not product_code:
            return PlacementResult(
                success=False,
                error_message=f"Unknown service type: {record.service}"
            )
        
        # Get format code
        format_type = self.normalise_format(record.format)
        format_code = self.get_format_code(format_type, dest_code)
        
        # Create order line
        order_line = SpringOrderLine(
            customer_number=self.CUSTOMER_NUMBER,
            customer_ref_1=self._po_number,
            customer_ref_2="",
            quote_ref="",
            count_sort=self.COUNT_SORT,
            pre_franked=self.PRE_FRANKED,
            product_code=product_code,
            nr_satchels=None,
            nr_bags=None,
            nr_boxes=None,
            nr_pallets=None,
            nr_trays=None,
            destination_code=dest_code,
            format_code=format_code,
            weightbreak_from=None,
            weightbreak_to=None,
            nr_items=record.items,
            weight_kg=record.weight,
        )
        
        self._order_lines.append(order_line)
        
        return PlacementResult(
            success=True,
            sheet_name="Orders",
            row=len(self._order_lines) + 1,
            items_col=17,
            weight_col=18,
        )
    
    def write_manifest(self, template_path: str, output_path: str) -> None:
        """
        Write accumulated order lines to the Spring manifest template.
        """
        wb = load_workbook(template_path)
        ws = wb['Orders']
        
        # Clear existing data rows (keep header row 1)
        for row in range(2, ws.max_row + 1):
            for col in range(1, 19):
                ws.cell(row=row, column=col).value = None
        
        # Group order lines by product code (1MI first, then 2MI)
        priority_lines = [line for line in self._order_lines if line.product_code == '1MI']
        economy_lines = [line for line in self._order_lines if line.product_code == '2MI']
        
        row_idx = 2  # Start writing from row 2 (row 1 is headers)
        
        # Write 1MI (Priority) order block
        for i, line in enumerate(priority_lines):
            # Order-level columns (A-L) only on first row of this product code block
            if i == 0:
                ws.cell(row=row_idx, column=1).value = line.customer_number
                ws.cell(row=row_idx, column=2).value = line.customer_ref_1
                ws.cell(row=row_idx, column=3).value = line.customer_ref_2
                ws.cell(row=row_idx, column=4).value = line.quote_ref
                ws.cell(row=row_idx, column=5).value = line.count_sort
                ws.cell(row=row_idx, column=6).value = line.pre_franked
                ws.cell(row=row_idx, column=7).value = line.product_code
                ws.cell(row=row_idx, column=8).value = line.nr_satchels
                ws.cell(row=row_idx, column=9).value = line.nr_bags
                ws.cell(row=row_idx, column=10).value = line.nr_boxes
                ws.cell(row=row_idx, column=11).value = 1  # Nr pallets - static
                ws.cell(row=row_idx, column=12).value = line.nr_trays
            
            # Order line columns (M-R) on every row
            ws.cell(row=row_idx, column=13).value = line.destination_code
            ws.cell(row=row_idx, column=14).value = line.format_code
            ws.cell(row=row_idx, column=15).value = line.weightbreak_from
            ws.cell(row=row_idx, column=16).value = line.weightbreak_to
            ws.cell(row=row_idx, column=17).value = line.nr_items
            ws.cell(row=row_idx, column=18).value = line.weight_kg
            row_idx += 1
        
        # Write 2MI (Economy) order block
        for i, line in enumerate(economy_lines):
            # Order-level columns (A-L) only on first row of this product code block
            if i == 0:
                ws.cell(row=row_idx, column=1).value = line.customer_number
                ws.cell(row=row_idx, column=2).value = line.customer_ref_1
                ws.cell(row=row_idx, column=3).value = line.customer_ref_2
                ws.cell(row=row_idx, column=4).value = line.quote_ref
                ws.cell(row=row_idx, column=5).value = line.count_sort
                ws.cell(row=row_idx, column=6).value = line.pre_franked
                ws.cell(row=row_idx, column=7).value = line.product_code
                ws.cell(row=row_idx, column=8).value = line.nr_satchels
                ws.cell(row=row_idx, column=9).value = line.nr_bags
                ws.cell(row=row_idx, column=10).value = line.nr_boxes
                ws.cell(row=row_idx, column=11).value = 1  # Nr pallets - static
                ws.cell(row=row_idx, column=12).value = line.nr_trays
            
            # Order line columns (M-R) on every row
            ws.cell(row=row_idx, column=13).value = line.destination_code
            ws.cell(row=row_idx, column=14).value = line.format_code
            ws.cell(row=row_idx, column=15).value = line.weightbreak_from
            ws.cell(row=row_idx, column=16).value = line.weightbreak_to
            ws.cell(row=row_idx, column=17).value = line.nr_items
            ws.cell(row=row_idx, column=18).value = line.weight_kg
            row_idx += 1
        
        # Delete Instructions and Product Combinations sheets
        for sheet_name in ['Instructions', 'Product Combinations']:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
        
        wb.save(output_path)
        wb.close()
    
    def clear_order_lines(self) -> None:
        """Clear accumulated order lines for fresh processing."""
        self._order_lines = []
        self._po_number = ""
    
    def get_order_lines(self) -> List[SpringOrderLine]:
        """Return accumulated order lines."""
        return self._order_lines
