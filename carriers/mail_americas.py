"""
Mail Americas/Africa carrier manifest handler.

Handles three distinct sheet structures:
- Mail Africa 2025: Weight-break based, Standard/Priority columns
- Mail Americas 2025: Weight-break based, Standard/Priority columns  
- Europe & ROW 2025: Format-based (Letters/Flats/Packets), no weight breaks
"""

from typing import Dict, Tuple, List, Optional
from openpyxl import load_workbook
from .base import BaseCarrier, ShipmentRecord, PlacementResult


class MailAmericasCarrier(BaseCarrier):
    """Handler for Mail Americas/Africa manifests."""
    
    carrier_name = "Mail Americas"
    template_filename = "Mail_America_Africa_2025.xlsx"
    
    # Service columns for Africa/Americas sheets (weight-break based)
    # Standard = Untracked Economy Mail, Priority = Untracked Priority Mail
    SERVICE_COLUMNS = {
        'Economy': {'items': 5, 'kg': 6},    # E, F
        'Priority': {'items': 7, 'kg': 8},   # G, H
    }
    
    # Format columns for Europe & ROW sheet
    FORMAT_COLUMNS = {
        'Letters': {'items': 4, 'kg': 5},   # D, E
        'Flats': {'items': 6, 'kg': 7},     # F, G
        'Packets': {'items': 8, 'kg': 9},   # H, I
    }
    
    def __init__(self):
        super().__init__()
        
        # Country name mappings (IST -> Manifest)
        self.country_mapping = {
            'Vietnam': 'Viet Nam',
            'Viet nam': 'Viet Nam',
            'UAE': 'United Arab Emirates',
            'U.A.E.': 'United Arab Emirates',
            'Czech Republic': 'Czech Rep',
            'South Korea': 'Korea',
            'Republic of Korea': 'Korea',
            'DRC': 'Congo, Democratic Republic.',
            'Democratic Republic of Congo': 'Congo, Democratic Republic.',
            'Republic of Congo': 'Congo, Republic of',
            "Cote d'Ivoire": 'Ivory Coast',
            "Côte d'Ivoire": 'Ivory Coast',
            'Bosnia': 'Bosnia & Herzegovina',
            'Serbia': 'Serbia & Montenegro',
            'Montenegro': 'Serbia & Montenegro',
            'Central African Republic': 'Central African Rep.',
        }
        
        # Caches for country->sheet/row mappings
        self._africa_index: Dict[str, List[Dict]] = {}
        self._americas_index: Dict[str, List[Dict]] = {}
        self._europe_row_index: Dict[str, List[Dict]] = {}
        self._index_built = False
    
    def _parse_weight_string(self, weight_str: str) -> Tuple[Optional[int], Optional[int]]:
        """Parse weight range string like '0 - 2000 grs' to (lower, upper)."""
        if not weight_str:
            return None, None
        cleaned = str(weight_str).replace('grs', '').replace('g', '').strip()
        parts = cleaned.split('-')
        if len(parts) == 2:
            try:
                return int(parts[0].strip()), int(parts[1].strip())
            except (ValueError, TypeError):
                pass
        return None, None
    
    def build_country_index(self, workbook) -> Dict[str, dict]:
        """
        Build index mapping countries to their sheet locations and weight breaks.
        
        Returns unified index for the base class interface, but internally
        maintains separate indices for each sheet type.
        """
        if self._index_built:
            return self._get_unified_index()
        
        # Build Mail Africa index
        ws = workbook['Mail Africa 2025']
        current_country = None
        for row in range(9, 67):
            country = ws.cell(row=row, column=2).value
            lower = ws.cell(row=row, column=3).value
            upper = ws.cell(row=row, column=4).value
            
            if country and str(country).strip() not in ['TOTALS:', 'Indicia Service ', 'GRAND TOTAL:']:
                current_country = str(country).strip()
                self._africa_index[current_country] = []
            
            if current_country and lower is not None and upper is not None:
                try:
                    self._africa_index[current_country].append({
                        'row': row,
                        'lower': int(float(lower)),
                        'upper': int(float(upper))
                    })
                except (ValueError, TypeError):
                    pass
        
        # Build Mail Americas index
        ws = workbook['Mail Americas 2025']
        current_country = None
        for row in range(9, 123):
            country = ws.cell(row=row, column=2).value
            lower = ws.cell(row=row, column=3).value
            upper = ws.cell(row=row, column=4).value
            
            if country and str(country).strip() not in ['TOTALS:', 'Indicia Service ', 'GRAND TOTAL:']:
                current_country = str(country).strip()
                self._americas_index[current_country] = []
            
            if current_country and lower is not None and upper is not None:
                try:
                    self._americas_index[current_country].append({
                        'row': row,
                        'lower': int(float(lower)),
                        'upper': int(float(upper))
                    })
                except (ValueError, TypeError):
                    pass
        
        # Build Europe & ROW index
        ws = workbook['Europe & ROW 2025']
        region_headers = ['AFRICA', 'AMERICAS', 'ASIA', 'EUROPE', 
                         'FAR EAST & AUSTRALASIA', 'MIDDLE EAST', 'OCEANIA']
        current_country = None
        
        for row in range(9, 123):
            country = ws.cell(row=row, column=2).value
            weight_str = ws.cell(row=row, column=3).value
            
            if country:
                country_str = str(country).strip()
                if country_str in region_headers or country_str in ['TOTALS:', 'GRAND TOTAL:']:
                    current_country = None
                    continue
                current_country = country_str
                self._europe_row_index[current_country] = []
            
            if current_country and weight_str:
                lower, upper = self._parse_weight_string(str(weight_str))
                if lower is not None:
                    self._europe_row_index[current_country].append({
                        'row': row,
                        'lower': lower,
                        'upper': upper
                    })
        
        self._index_built = True
        return self._get_unified_index()
    
    def _get_unified_index(self) -> Dict[str, dict]:
        """Create unified index for base class compatibility."""
        index = {}
        
        # Add Africa countries
        for country, breaks in self._africa_index.items():
            index[country] = {
                'Priority': {'sheet': 'Mail Africa 2025', 'breaks': breaks, 'type': 'weight_break'},
                'Economy': {'sheet': 'Mail Africa 2025', 'breaks': breaks, 'type': 'weight_break'}
            }
        
        # Add Americas countries
        for country, breaks in self._americas_index.items():
            index[country] = {
                'Priority': {'sheet': 'Mail Americas 2025', 'breaks': breaks, 'type': 'weight_break'},
                'Economy': {'sheet': 'Mail Americas 2025', 'breaks': breaks, 'type': 'weight_break'}
            }
        
        # Add Europe & ROW countries (format-based, not service-based)
        for country, breaks in self._europe_row_index.items():
            index[country] = {
                'Priority': {'sheet': 'Europe & ROW 2025', 'breaks': breaks, 'type': 'format_based'},
                'Economy': {'sheet': 'Europe & ROW 2025', 'breaks': breaks, 'type': 'format_based'}
            }
        
        return index
    
    def _find_country_sheet(self, country: str) -> Tuple[Optional[str], Optional[List[Dict]]]:
        """Find which sheet a country belongs to."""
        mapped = self.map_country(country) or country
        
        # Try exact match first
        if mapped in self._africa_index:
            return 'Mail Africa 2025', self._africa_index[mapped]
        if mapped in self._americas_index:
            return 'Mail Americas 2025', self._americas_index[mapped]
        if mapped in self._europe_row_index:
            return 'Europe & ROW 2025', self._europe_row_index[mapped]
        
        # Try case-insensitive and whitespace-normalized match
        for idx, sheet_name in [(self._africa_index, 'Mail Africa 2025'),
                                (self._americas_index, 'Mail Americas 2025'),
                                (self._europe_row_index, 'Europe & ROW 2025')]:
            for template_country in idx.keys():
                if template_country.strip().lower() == mapped.strip().lower():
                    return sheet_name, idx[template_country]
        
        return None, None
    
    def _find_weight_break_row(self, avg_weight_kg: float, breaks: List[Dict]) -> Optional[int]:
        """Find manifest row matching the average weight."""
        if not breaks:
            return None
        
        avg_weight_g = avg_weight_kg * 1000
        
        for wb in breaks:
            if wb['lower'] <= avg_weight_g <= wb['upper']:
                return wb['row']
        
        # Allow small tolerance for edge cases
        for wb in breaks:
            if wb['lower'] - 1 <= avg_weight_g <= wb['upper'] + 1:
                return wb['row']
        
        return None
    
    def get_cell_positions(self, country_info: dict, format_type: str) -> Tuple[int, int]:
        """Get columns for items and weight."""
        if country_info['type'] == 'format_based':
            # Europe & ROW - use format columns
            fmt = format_type if format_type in self.FORMAT_COLUMNS else 'Flats'
            cols = self.FORMAT_COLUMNS[fmt]
            return cols['items'], cols['kg']
        else:
            # This shouldn't be called for weight-break sheets
            # as we handle them differently in place_record
            raise ValueError("Weight-break sheets use custom placement logic")
    
    def set_metadata(self, workbook, po_number: str, shipment_date: str) -> None:
        """Set PO and date on all three sheets."""
        from datetime import datetime
        
        # Format date as DD/MM/YYYY
        if isinstance(shipment_date, str) and shipment_date:
            try:
                dt = datetime.strptime(shipment_date, "%Y-%m-%d")
                date_formatted = dt.strftime("%d/%m/%Y")
            except ValueError:
                date_formatted = shipment_date
        else:
            date_formatted = datetime.now().strftime("%d/%m/%Y")
        
        for sheet_name in ['Mail Africa 2025', 'Mail Americas 2025', 'Europe & ROW 2025']:
            ws = workbook[sheet_name]
            
            # PO in merged cell E5:F5
            ws['E5'] = f'REF: {po_number}'
            
            # Date position differs for Europe & ROW
            if sheet_name == 'Europe & ROW 2025':
                ws['H5'] = f'DATE: {date_formatted}'
            else:
                ws['G5'] = f'DATE: {date_formatted}'
    
    def place_record(self, workbook, record: ShipmentRecord, country_index: dict) -> PlacementResult:
        """
        Place a shipment record into the manifest.
        
        Overrides base class to handle:
        - Weight-break matching for Africa/Americas
        - Format-based placement for Europe & ROW
        """
        # Find country's sheet and weight breaks
        sheet_name, breaks = self._find_country_sheet(record.country)
        
        if not sheet_name:
            return PlacementResult(
                success=False,
                error_message=f"Country not found in manifest: {record.country}"
            )
        
        ws = workbook[sheet_name]
        service = self.normalise_service(record.service)
        format_type = self.normalise_format(record.format)
        
        if sheet_name == 'Europe & ROW 2025':
            # Format-based placement - use first row for country
            if not breaks:
                return PlacementResult(
                    success=False,
                    error_message=f"No row mapping for {record.country} in Europe & ROW"
                )
            
            target_row = breaks[0]['row']
            cols = self.FORMAT_COLUMNS.get(format_type, self.FORMAT_COLUMNS['Flats'])
            
            self._add_to_cell(ws, target_row, cols['items'], record.items)
            self._add_to_cell(ws, target_row, cols['kg'], record.weight)
            
        else:
            # Weight-break based placement (Africa/Americas)
            cols = self.SERVICE_COLUMNS.get(service, self.SERVICE_COLUMNS['Economy'])
            
            # Calculate average weight per item for weight break matching
            if record.items > 0:
                avg_weight_kg = record.weight / record.items
            else:
                avg_weight_kg = 0
            
            target_row = self._find_weight_break_row(avg_weight_kg, breaks)
            
            if not target_row:
                return PlacementResult(
                    success=False,
                    error_message=f"No matching weight break for {record.country} (avg: {avg_weight_kg:.3f}kg)"
                )
            
            self._add_to_cell(ws, target_row, cols['items'], record.items)
            self._add_to_cell(ws, target_row, cols['kg'], record.weight)
        
        return PlacementResult(
            success=True,
            sheet_name=sheet_name,
            row=target_row,
            items_col=cols['items'],
            weight_col=cols['kg']
        )
    
    def _add_to_cell(self, ws, row: int, col: int, value: float) -> None:
        """Add value to existing cell value."""
        current = ws.cell(row=row, column=col).value
        if current is None or current == '':
            current = 0
        try:
            current = float(current)
        except (ValueError, TypeError):
            current = 0
        ws.cell(row=row, column=col).value = current + value
