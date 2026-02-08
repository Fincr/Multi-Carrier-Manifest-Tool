"""
Multi-Carrier Manifest Population Tool
======================================
GUI for populating carrier manifests from internal carrier sheets.

Supports:
- Asendia UK Business Mail
- PostNord Business Mail
- Spring Global Delivery Solutions
- Landmark Global
- Deutsche Post
- Air Business
- Mail Americas

Usage:
    python gui.py
"""

__version__ = "1.4.3"
__author__ = "Finlay Crawley"

import sys
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk
from datetime import datetime
import asyncio

# Add parent directory to path for imports when running directly
if __name__ == "__main__":
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from core.engine import ManifestEngine
from core.config import get_config, save_config, AppConfig, get_available_printers
from core.credentials import get_landmark_credentials
from pre_alerts.pre_alert_tab import PreAlertTab
from pre_alerts.config_manager import load_pre_alert_config


def get_portal_timeout_ms() -> int:
    """Get portal timeout from config."""
    return get_config().portal_timeout_ms


def get_portal_retry_count() -> int:
    """Get portal retry count from config."""
    return get_config().portal_retry_count


def print_pdf_file(filepath: str, printer_name: str = None, close_after: bool = True, close_delay: int = None) -> tuple[bool, str]:
    """
    Print a PDF file to the specified printer.
    
    Tries multiple methods in order of preference:
    1. SumatraPDF portable (bundled in tools folder - fast, silent, supports simplex)
    2. Adobe Acrobat/Reader (fallback if SumatraPDF not available)
    3. Windows shell print verb (last resort - may open dialogs)
    
    Args:
        filepath: Full path to the PDF file
        printer_name: Network printer name (optional, uses default if not specified)
        close_after: Whether to close the PDF app after printing (default True)
        close_delay: Seconds to wait before closing the app (default 7)
    
    Returns:
        (success: bool, message: str)
    """
    try:
        import subprocess
        import time
        
        # Use config default if close_delay not specified
        if close_delay is None:
            close_delay = get_config().pdf_close_delay_seconds
        
        # Check if file exists
        if not os.path.exists(filepath):
            return False, f"PDF file not found: {filepath}"
        
        # Get the app directory for locating bundled tools
        app_dir = os.path.dirname(os.path.abspath(__file__))
        
        # =================================================================
        # METHOD 1: Try SumatraPDF (portable, fast, silent, supports simplex)
        # =================================================================
        # Look for SumatraPDF in multiple locations
        sumatra_paths = [
            os.path.join(app_dir, "tools", "SumatraPDF.exe"),
            os.path.join(app_dir, "SumatraPDF.exe"),
            r"C:\Program Files\SumatraPDF\SumatraPDF.exe",
            r"C:\Program Files (x86)\SumatraPDF\SumatraPDF.exe",
            # Also check common portable locations
            os.path.join(os.environ.get('LOCALAPPDATA', ''), 'SumatraPDF', 'SumatraPDF.exe'),
        ]
        
        sumatra_exe = None
        for path in sumatra_paths:
            if path and os.path.exists(path):
                sumatra_exe = path
                break
        
        if sumatra_exe:
            # SumatraPDF command line:
            # -print-to "printer name" : print to specific printer
            # -print-to-default : print to default printer  
            # -silent : don't show any dialogs or windows
            # -print-settings "simplex" : force single-sided printing
            # -exit-when-done : close after printing (implicit with -print-to)
            if printer_name:
                cmd = [sumatra_exe, '-print-to', printer_name, '-print-settings', 'simplex', '-silent', filepath]
            else:
                cmd = [sumatra_exe, '-print-to-default', '-print-settings', 'simplex', '-silent', filepath]
            
            # Run and wait briefly for print job to spool
            proc = subprocess.Popen(cmd)
            
            # SumatraPDF exits automatically after printing, but give it time
            def wait_and_cleanup():
                try:
                    proc.wait(timeout=30)  # Wait up to 30 seconds
                except subprocess.TimeoutExpired:
                    proc.terminate()
            
            threading.Thread(target=wait_and_cleanup, daemon=True).start()
            
            return True, "Sent to printer via SumatraPDF (simplex)"
        
        # =================================================================
        # METHOD 2: Try Adobe Acrobat/Reader (fallback)
        # =================================================================
        adobe_paths = [
            r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
            r"C:\Program Files (x86)\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
            r"C:\Program Files\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
            r"C:\Program Files (x86)\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
        ]
        
        adobe_exe = None
        for path in adobe_paths:
            if os.path.exists(path):
                adobe_exe = path
                break
        
        if adobe_exe:
            # Use Adobe Reader to print silently
            # /t switch: print to specified printer and exit
            # /h switch: minimized
            # Note: Adobe uses printer's default duplex settings
            if printer_name:
                cmd = [adobe_exe, '/t', filepath, printer_name]
            else:
                cmd = [adobe_exe, '/p', '/h', filepath]
            proc = subprocess.Popen(cmd)
            
            # Close Adobe after printing if requested
            if close_after:
                def close_adobe_later():
                    time.sleep(close_delay)
                    try:
                        proc.terminate()
                    except Exception:
                        pass
                    try:
                        subprocess.run(['taskkill', '/F', '/IM', 'Acrobat.exe'],
                                     capture_output=True, timeout=5)
                        subprocess.run(['taskkill', '/F', '/IM', 'AcroRd32.exe'],
                                     capture_output=True, timeout=5)
                    except Exception:
                        pass

                threading.Thread(target=close_adobe_later, daemon=True).start()
            
            return True, "Sent to printer via Adobe (uses printer duplex settings)"
        
        # =================================================================
        # METHOD 3: Windows shell print verb (last resort)
        # =================================================================
        # This opens whatever app is associated with PDFs and sends print command
        # May show print dialogs depending on the default PDF app
        import ctypes
        result = ctypes.windll.shell32.ShellExecuteW(
            None, "print", filepath, None, None, 0
        )
        
        if result > 32:
            # Try to close common PDF viewers after printing
            if close_after:
                def close_pdf_viewers_later():
                    time.sleep(close_delay)
                    # Try to close various PDF viewers that might have opened
                    viewers_to_close = [
                        'Acrobat.exe', 'AcroRd32.exe',  # Adobe
                        'SumatraPDF.exe',                # Sumatra
                        'FoxitPDFReader.exe', 'FoxitReader.exe',  # Foxit
                        'PDFXEdit.exe',                  # PDF-XChange
                    ]
                    for viewer in viewers_to_close:
                        try:
                            subprocess.run(['taskkill', '/F', '/IM', viewer],
                                         capture_output=True, timeout=5)
                        except Exception:
                            pass
                
                threading.Thread(target=close_pdf_viewers_later, daemon=True).start()
            
            return True, "Sent to default printer via Windows shell (may require interaction)"
        else:
            return False, f"ShellExecute failed with code {result}"
            
    except Exception as e:
        return False, f"Print failed: {str(e)}"


def print_excel_workbook(filepath: str, printer_name: str = None) -> tuple[bool, str]:
    """
    Print an Excel workbook to the specified printer.
    
    Uses Windows COM automation via win32com.
    Falls back to default printer if printer_name not specified.
    Sets "Fit All Columns on One Page" for each sheet before printing.
    
    Args:
        filepath: Full path to the Excel file
        printer_name: Network printer name (e.g., '\\\\print01.citipost.co.uk\\KT02')
                     If None, uses Windows default printer
    
    Returns:
        (success: bool, message: str)
    """
    try:
        import win32com.client
        import pythoncom
    except ImportError:
        return False, "pywin32 not installed. Run: pip install pywin32"
    
    excel = None
    wb = None
    
    try:
        # Initialize COM for this thread
        pythoncom.CoInitialize()
        
        # Create Excel instance
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open workbook
        wb = excel.Workbooks.Open(filepath)
        
        # Set "Fit All Columns on One Page" for each sheet
        for sheet in wb.Sheets:
            sheet.PageSetup.Zoom = False  # Disable fixed zoom to allow fit-to-page
            sheet.PageSetup.FitToPagesWide = 1  # Fit all columns to 1 page wide
            sheet.PageSetup.FitToPagesTall = False  # Don't constrain rows (allow multiple pages tall)
        
        # Print entire workbook
        if printer_name:
            # Print to specific printer
            wb.PrintOut(ActivePrinter=printer_name)
        else:
            # Print to default printer
            wb.PrintOut()
        
        return True, f"Sent to printer: {printer_name or 'default'}"
        
    except Exception as e:
        return False, f"Print failed: {str(e)}"
        
    finally:
        # Cleanup
        if wb:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def run_spring_upload(file_path: str, po_number: str = "", output_dir: str = "", auto_print: bool = True, log_callback=None) -> tuple[bool, str, bool]:
    """
    Synchronous wrapper for the robust Spring portal upload.
    
    Uses the new robust portal automation with comprehensive error handling,
    per-stage retry logic, and graceful degradation for unreliable portal behaviour.
    
    Known portal issues handled:
    - Post-login hang (page doesn't fully load after authentication)
    - "Upload Multiple Orders" button not found (dynamic UI loading)
    - "Unexpected error" after CSV upload
    - "Unexpected error" when downloading PDF manifest
    
    Args:
        file_path: Path to the upload file (Excel)
        po_number: PO/Customer ref to find after upload (for printing)
        output_dir: Directory to save the downloaded PDF manifest
        auto_print: Whether to print the downloaded PDF
        log_callback: Optional callback for logging messages
        
    Returns:
        (success: bool, message: str, pdf_downloaded: bool)
        pdf_downloaded indicates whether the manifest PDF was successfully downloaded/printed
    """
    try:
        from carriers.spring_portal import run_spring_upload_robust
        return run_spring_upload_robust(file_path, po_number, output_dir, auto_print, log_callback)
    except ImportError as e:
        # Fallback error if module not found
        return False, f"Spring portal module not found: {str(e)}", False
    except Exception as e:
        return False, f"Upload error: {str(e)}", False


async def _upload_to_landmark_portal_impl(file_paths: list, po_number: str = "", output_dir: str = "", auto_print: bool = True, log_callback=None) -> tuple[bool, str, bool]:
    """
    Internal implementation of Landmark portal upload.
    Called by upload_to_landmark_portal with retry logic.
    
    Returns:
        (success: bool, message: str, pdf_downloaded: bool)
        pdf_downloaded indicates whether manifest PDFs were successfully downloaded/printed
    """
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        return False, "Playwright not installed. Run: pip install playwright && playwright install chromium", False
    
    # Load credentials from environment/.env
    creds = get_landmark_credentials()
    if not creds.is_valid():
        return False, "Landmark credentials not configured. Set LANDMARK_EMAIL and LANDMARK_PASSWORD in .env file.", False
    
    USERNAME = creds.email
    PASSWORD = creds.password
    LOGIN_URL = "https://www.bpost.be/portal/goLogin?oss_language=EN"
    
    def log(msg):
        if log_callback:
            log_callback(msg)
    
    downloaded_files = []
    
    try:
        async with async_playwright() as p:
            # Launch browser
            log("  Launching browser...")
            browser = await p.chromium.launch(headless=False)
            context = await browser.new_context(accept_downloads=True)
            page = await context.new_page()
            
            try:
                # Navigate to login page
                log("  Navigating to Landmark portal...")
                await page.goto(LOGIN_URL, wait_until="networkidle", timeout=get_portal_timeout_ms())
                
                # Handle cookie consent banner if present
                log("  Checking for cookie consent...")
                cookie_selectors = [
                    'button:has-text("Accept")',
                    'button:has-text("Accept all")',
                    'button:has-text("Accept All")',
                    'button:has-text("Accepteer")',
                    'button:has-text("Akkoord")',
                    'button:has-text("OK")',
                    'button:has-text("I agree")',
                    'button:has-text("Agree")',
                    'button[id*="accept"]',
                    'button[id*="cookie"]',
                    'a:has-text("Accept")',
                    '#onetrust-accept-btn-handler',
                    '.cookie-accept',
                    '[data-testid="cookie-accept"]',
                ]
                
                for selector in cookie_selectors:
                    try:
                        element = page.locator(selector).first
                        if await element.is_visible(timeout=2000):
                            await element.click()
                            log("    ✓ Cookie consent accepted")
                            await page.wait_for_timeout(1000)
                            break
                    except Exception:
                        continue

                # Enter credentials
                log("  Entering credentials...")
                await page.wait_for_selector('input[name="username"], input[type="text"], input[id*="user"], input[name="j_username"]', timeout=get_portal_timeout_ms())
                
                # Find and fill username field
                username_selectors = ['input[name="j_username"]', 'input[name="username"]', 'input[type="text"]', 'input[id*="user"]', 'input[name="user"]']
                for selector in username_selectors:
                    try:
                        element = page.locator(selector).first
                        if await element.is_visible(timeout=1000):
                            await element.fill(USERNAME)
                            log("    ✓ Username entered")
                            break
                    except Exception:
                        continue
                
                # Find and fill password field
                password_selectors = ['input[name="j_password"]', 'input[type="password"]']
                for selector in password_selectors:
                    try:
                        element = page.locator(selector).first
                        if await element.is_visible(timeout=1000):
                            await element.fill(PASSWORD)
                            log("    ✓ Password entered")
                            break
                    except Exception:
                        continue
                
                # Click Sign In button
                log("  Signing in...")
                sign_in_selectors = [
                    'button:has-text("Sign In")',
                    'button:has-text("Sign in")',
                    'button:has-text("Login")',
                    'button:has-text("Log in")',
                    'button:has-text("Aanmelden")',
                    'input[type="submit"]',
                    'button[type="submit"]',
                ]
                
                for selector in sign_in_selectors:
                    try:
                        element = page.locator(selector).first
                        if await element.is_visible(timeout=1000):
                            await element.click()
                            break
                    except Exception:
                        continue
                
                await page.wait_for_load_state("networkidle", timeout=get_portal_timeout_ms())
                await page.wait_for_timeout(3000)
                
                # Check for cookie consent again after login (some sites show it again)
                for selector in cookie_selectors:
                    try:
                        element = page.locator(selector).first
                        if await element.is_visible(timeout=1000):
                            await element.click()
                            log("    ✓ Cookie consent accepted (post-login)")
                            await page.wait_for_timeout(1000)
                            break
                    except Exception:
                        continue

                log("  ✓ Logged in successfully")
                
                # Click "e-Shipper bpost international"
                log("  Navigating to e-Shipper...")
                eshipper_selectors = [
                    'text="e-Shipper bpost international"',
                    'text="e-Shipper"',
                    'a:has-text("e-Shipper")',
                    'a:has-text("Shipper")',
                    '[href*="shipper"]',
                    '[href*="eshipper"]',
                ]
                
                clicked = False
                for selector in eshipper_selectors:
                    try:
                        element = page.locator(selector).first
                        if await element.is_visible(timeout=3000):
                            await element.click()
                            clicked = True
                            log("    ✓ Found e-Shipper link")
                            break
                    except Exception:
                        continue
                
                if not clicked:
                    screenshot_path = os.path.join(output_dir, "landmark_debug_eshipper.png")
                    await page.screenshot(path=screenshot_path)
                    log(f"  ⚠ Could not find e-Shipper link. Screenshot saved: {screenshot_path}")
                    await browser.close()
                    return False, "Could not find e-Shipper link. Check landmark_debug_eshipper.png", False
                
                await page.wait_for_load_state("networkidle", timeout=get_portal_timeout_ms())
                await page.wait_for_timeout(2000)
                
                # Process each file
                for file_idx, file_path in enumerate(file_paths):
                    filename = os.path.basename(file_path)
                    service_type = "Economy" if "Economy" in filename else "Priority"
                    log(f"\n  Processing {service_type} file ({file_idx + 1}/{len(file_paths)}): {filename}")
                    
                    # Click "Upload deposit csv" on left side
                    log("    Clicking 'Upload deposit csv'...")
                    upload_csv_selectors = [
                        'text="Upload deposit csv"',
                        'text="Upload deposit"',
                        'text="Upload CSV"',
                        'a:has-text("Upload deposit")',
                        'a:has-text("Upload")',
                        '[href*="upload"]',
                        '[href*="csv"]',
                    ]
                    
                    clicked = False
                    for selector in upload_csv_selectors:
                        try:
                            element = page.locator(selector).first
                            if await element.is_visible(timeout=3000):
                                await element.click()
                                clicked = True
                                log("      ✓ Found upload link")
                                break
                        except Exception:
                            continue
                    
                    if not clicked:
                        screenshot_path = os.path.join(output_dir, f"landmark_debug_upload_{service_type}.png")
                        await page.screenshot(path=screenshot_path)
                        log("    ⚠ Could not find 'Upload deposit csv' link. Screenshot saved.")
                        continue
                    
                    await page.wait_for_load_state("networkidle", timeout=get_portal_timeout_ms())
                    await page.wait_for_timeout(1000)
                    
                    # Click "Choose File" and select file
                    log("    Selecting file...")
                    try:
                        file_input = page.locator('input[type="file"]').first
                        await file_input.set_input_files(file_path)
                        log("      ✓ File selected")
                    except Exception as e:
                        log(f"    ⚠ Could not select file: {e}")
                        continue
                    
                    await page.wait_for_timeout(1000)
                    
                    # Click "Upload" button
                    log("    Clicking 'Upload'...")
                    upload_btn_selectors = [
                        'button:has-text("Upload")',
                        'input[value="Upload"]',
                        'input[type="submit"]',
                        'button[type="submit"]',
                    ]
                    
                    clicked = False
                    for selector in upload_btn_selectors:
                        try:
                            element = page.locator(selector).first
                            if await element.is_visible(timeout=2000):
                                await element.click()
                                clicked = True
                                log("      ✓ Upload button clicked")
                                break
                        except Exception:
                            continue
                    
                    if not clicked:
                        log("    ⚠ Could not find Upload button")
                        continue
                    
                    await page.wait_for_load_state("networkidle", timeout=get_portal_timeout_ms())
                    await page.wait_for_timeout(2000)
                    
                    # Check for specific validation errors (be more precise than just "error" in page)
                    page_text = await page.inner_text('body')
                    
                    # Look for specific error indicators
                    error_indicators = [
                        'invalid file',
                        'validation error',
                        'file format',
                        'incorrect format',
                        'upload failed',
                        'error uploading',
                    ]
                    
                    has_error = any(indicator in page_text.lower() for indicator in error_indicators)
                    
                    if has_error:
                        screenshot_path = os.path.join(output_dir, f"landmark_error_{service_type}.png")
                        await page.screenshot(path=screenshot_path)
                        log(f"    ⚠ Validation error detected. Screenshot saved: {screenshot_path}")
                        continue
                    
                    # Click "Final acceptance of input"
                    log("    Clicking 'Final acceptance of input'...")
                    final_selectors = [
                        'text="Final acceptance of input"',
                        'text="Final acceptance"',
                        'button:has-text("Final acceptance")',
                        'a:has-text("Final acceptance")',
                        'input[value*="Final"]',
                        'button:has-text("Confirm")',
                        'button:has-text("Submit")',
                    ]
                    
                    clicked = False
                    for selector in final_selectors:
                        try:
                            element = page.locator(selector).first
                            if await element.is_visible(timeout=5000):
                                # Try multiple approaches to capture the PDF
                                
                                # Approach 1: Listen for new popup/tab that might contain PDF
                                try:
                                    async with context.expect_page(timeout=10000) as new_page_info:
                                        await element.click()
                                        clicked = True
                                    
                                    new_page = await new_page_info.value
                                    await new_page.wait_for_load_state("networkidle", timeout=15000)
                                    
                                    # Check if it's a PDF URL
                                    new_url = new_page.url
                                    log(f"      New page opened: {new_url[:80]}...")
                                    
                                    if '.pdf' in new_url.lower() or 'pdf' in new_url.lower():
                                        # It's a PDF - save it
                                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                        downloaded_filename = f"Landmark_{service_type}_{po_number}_{timestamp}.pdf"
                                        downloaded_path = os.path.join(output_dir, downloaded_filename)
                                        
                                        # Try to download the PDF content
                                        try:
                                            # Use page to download
                                            pdf_response = await new_page.context.request.get(new_url)
                                            pdf_bytes = await pdf_response.body()
                                            
                                            with open(downloaded_path, 'wb') as f:
                                                f.write(pdf_bytes)
                                            
                                            downloaded_files.append(downloaded_path)
                                            log(f"    ✓ Downloaded: {downloaded_filename}")
                                        except Exception as pdf_err:
                                            log(f"      Could not save PDF directly: {pdf_err}")
                                            # Try printing the page to PDF as fallback
                                            try:
                                                await new_page.pdf(path=downloaded_path)
                                                downloaded_files.append(downloaded_path)
                                                log(f"    ✓ Saved as PDF: {downloaded_filename}")
                                            except Exception as print_err:
                                                log(f"      Could not print to PDF: {print_err}")
                                    
                                    await new_page.close()
                                    break
                                    
                                except Exception:
                                    # No popup - try download approach
                                    pass
                                
                                # Approach 2: Try expecting a download
                                if not clicked:
                                    try:
                                        async with page.expect_download(timeout=10000) as download_info:
                                            await element.click()
                                            clicked = True
                                        
                                        download = await download_info.value
                                        
                                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                        downloaded_filename = f"Landmark_{service_type}_{po_number}_{timestamp}.pdf"
                                        downloaded_path = os.path.join(output_dir, downloaded_filename)
                                        await download.save_as(downloaded_path)
                                        downloaded_files.append(downloaded_path)
                                        log(f"    ✓ Downloaded: {downloaded_filename}")
                                        break
                                    except Exception:
                                        # Click without download capture
                                        if not clicked:
                                            await element.click()
                                            clicked = True
                                        log("      Note: PDF may have opened in browser")
                                
                                break
                        except Exception:
                            continue
                    
                    if not clicked:
                        screenshot_path = os.path.join(output_dir, f"landmark_debug_final_{service_type}.png")
                        await page.screenshot(path=screenshot_path)
                        log("    ⚠ Could not find 'Final acceptance' button. Screenshot saved.")
                    
                    await page.wait_for_timeout(2000)
                
                await browser.close()
                
                # Print downloaded PDFs if auto_print enabled
                if auto_print and downloaded_files:
                    log("\n  Printing downloaded manifests...")
                    for pdf_path in downloaded_files:
                        print_success, print_msg = print_pdf_file(pdf_path)
                        if print_success:
                            log(f"    ✓ {os.path.basename(pdf_path)}: {print_msg}")
                        else:
                            log(f"    ⚠ {os.path.basename(pdf_path)}: {print_msg}")
                
                if downloaded_files:
                    return True, f"Successfully uploaded {len(file_paths)} file(s) and downloaded {len(downloaded_files)} manifest(s)", True
                else:
                    return True, f"Uploaded {len(file_paths)} file(s) but no manifests were downloaded", False
                
            except Exception as e:
                try:
                    await browser.close()
                except Exception:
                    pass
                raise e

    except Exception as e:
        return False, f"Upload failed: {str(e)}", False


async def upload_to_landmark_portal(file_paths: list, po_number: str = "", output_dir: str = "", auto_print: bool = True, log_callback=None) -> tuple[bool, str, bool]:
    """
    Upload CSV files to the Landmark Global portal using Playwright.
    Handles both Economy and Priority files sequentially.
    
    Includes automatic retry on timeout errors.
    
    Args:
        file_paths: List of CSV file paths to upload (Economy and/or Priority)
        po_number: PO number for naming downloaded files
        output_dir: Directory to save the downloaded PDF manifests
        auto_print: Whether to print the downloaded PDFs
        log_callback: Optional callback for logging messages
        
    Returns:
        (success: bool, message: str, pdf_downloaded: bool)
        pdf_downloaded indicates whether manifest PDFs were successfully downloaded/printed
    """
    def log(msg):
        if log_callback:
            log_callback(msg)
    
    last_error = None
    
    for attempt in range(get_portal_retry_count() + 1):
        if attempt > 0:
            log(f"\n  ⟳ Retry attempt {attempt} of {get_portal_retry_count()}...")
        
        success, message, pdf_downloaded = await _upload_to_landmark_portal_impl(
            file_paths, po_number, output_dir, auto_print, log_callback
        )
        
        if success:
            return success, message, pdf_downloaded
        
        # Check if this is a timeout error worth retrying
        last_error = message
        if "Timeout" in message or "timeout" in message:
            if attempt < get_portal_retry_count():
                log("  ⚠ Timeout occurred, will retry...")
                continue
        else:
            # Non-timeout error, don't retry
            break
    
    return False, last_error or "Upload failed after retries", False


def run_landmark_upload(file_paths: list, po_number: str = "", output_dir: str = "", auto_print: bool = True, log_callback=None) -> tuple[bool, str, bool]:
    """
    Synchronous wrapper for the async Landmark upload function.
    
    Args:
        file_paths: List of CSV file paths to upload
        po_number: PO number for naming downloaded files
        output_dir: Directory to save the downloaded PDF manifests
        auto_print: Whether to print the downloaded PDFs
        log_callback: Optional callback for logging messages
        
    Returns:
        (success: bool, message: str, pdf_downloaded: bool)
        pdf_downloaded indicates whether manifest PDFs were successfully downloaded/printed
    """
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            return loop.run_until_complete(upload_to_landmark_portal(file_paths, po_number, output_dir, auto_print, log_callback))
        finally:
            loop.close()
    except Exception as e:
        return False, f"Upload error: {str(e)}", False


class SettingsDialog:
    """Settings dialog for configuring application options."""
    
    def __init__(self, parent, config: AppConfig):
        self.result = None
        self.config = config
        
        # Create dialog window
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Settings")
        self.dialog.geometry("500x450")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center on parent
        self.dialog.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (500 // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (450 // 2)
        self.dialog.geometry(f"+{x}+{y}")
        
        self._create_widgets()
        
    def _create_widgets(self):
        """Build the settings dialog UI."""
        main_frame = ttk.Frame(self.dialog, padding="15")
        main_frame.pack(fill='both', expand=True)
        
        # === Printer Settings ===
        printer_frame = ttk.LabelFrame(main_frame, text="Printer Settings", padding="10")
        printer_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(printer_frame, text="Printer:").grid(row=0, column=0, sticky='w', pady=5)
        
        # Printer dropdown
        self.printer_var = tk.StringVar(value=self.config.printer_name)
        printers = get_available_printers()
        
        # Add current printer if not in list
        if self.config.printer_name and self.config.printer_name not in printers:
            printers.insert(0, self.config.printer_name)
        
        if not printers:
            printers = [self.config.printer_name or "No printers found"]
        
        self.printer_combo = ttk.Combobox(
            printer_frame, 
            textvariable=self.printer_var,
            values=printers,
            width=50
        )
        self.printer_combo.grid(row=0, column=1, sticky='ew', padx=(10, 0))
        printer_frame.columnconfigure(1, weight=1)
        
        # === Portal Settings ===
        portal_frame = ttk.LabelFrame(main_frame, text="Portal Settings", padding="10")
        portal_frame.pack(fill='x', pady=(0, 10))
        
        # Timeout
        ttk.Label(portal_frame, text="Timeout (seconds):").grid(row=0, column=0, sticky='w', pady=5)
        self.timeout_var = tk.StringVar(value=str(self.config.portal_timeout_ms // 1000))
        timeout_spin = ttk.Spinbox(
            portal_frame,
            from_=5,
            to=120,
            width=10,
            textvariable=self.timeout_var
        )
        timeout_spin.grid(row=0, column=1, sticky='w', padx=(10, 0))
        ttk.Label(portal_frame, text="(how long to wait for portal pages)", foreground='gray').grid(
            row=0, column=2, sticky='w', padx=(10, 0)
        )
        
        # Retry count
        ttk.Label(portal_frame, text="Retry count:").grid(row=1, column=0, sticky='w', pady=5)
        self.retry_var = tk.StringVar(value=str(self.config.portal_retry_count))
        retry_spin = ttk.Spinbox(
            portal_frame,
            from_=0,
            to=5,
            width=10,
            textvariable=self.retry_var
        )
        retry_spin.grid(row=1, column=1, sticky='w', padx=(10, 0))
        ttk.Label(portal_frame, text="(retries on timeout errors)", foreground='gray').grid(
            row=1, column=2, sticky='w', padx=(10, 0)
        )
        
        # === Print Settings ===
        print_frame = ttk.LabelFrame(main_frame, text="Print Settings", padding="10")
        print_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(print_frame, text="PDF close delay (seconds):").grid(row=0, column=0, sticky='w', pady=5)
        self.close_delay_var = tk.StringVar(value=str(self.config.pdf_close_delay_seconds))
        close_delay_spin = ttk.Spinbox(
            print_frame,
            from_=1,
            to=30,
            width=10,
            textvariable=self.close_delay_var
        )
        close_delay_spin.grid(row=0, column=1, sticky='w', padx=(10, 0))
        ttk.Label(print_frame, text="(wait before closing PDF viewer)", foreground='gray').grid(
            row=0, column=2, sticky='w', padx=(10, 0)
        )
        
        # === Processing Settings ===
        proc_frame = ttk.LabelFrame(main_frame, text="Processing Settings", padding="10")
        proc_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(proc_frame, text="Max errors before stop:").grid(row=0, column=0, sticky='w', pady=5)
        self.max_errors_var = tk.StringVar(value=str(self.config.max_errors_before_stop))
        max_errors_spin = ttk.Spinbox(
            proc_frame,
            from_=1,
            to=50,
            width=10,
            textvariable=self.max_errors_var
        )
        max_errors_spin.grid(row=0, column=1, sticky='w', padx=(10, 0))
        ttk.Label(proc_frame, text="(stop processing after this many errors)", foreground='gray').grid(
            row=0, column=2, sticky='w', padx=(10, 0)
        )
        
        # === Buttons ===
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x', pady=(15, 0))
        
        ttk.Button(button_frame, text="Save", command=self._save).pack(side='right', padx=(5, 0))
        ttk.Button(button_frame, text="Cancel", command=self._cancel).pack(side='right')
        ttk.Button(button_frame, text="Reset to Defaults", command=self._reset_defaults).pack(side='left')
        
    def _save(self):
        """Save settings and close dialog."""
        try:
            # Validate and collect values
            timeout_seconds = int(self.timeout_var.get())
            retry_count = int(self.retry_var.get())
            close_delay = int(self.close_delay_var.get())
            max_errors = int(self.max_errors_var.get())
            
            # Update config
            self.config.printer_name = self.printer_var.get()
            self.config.portal_timeout_ms = timeout_seconds * 1000
            self.config.portal_retry_count = retry_count
            self.config.pdf_close_delay_seconds = close_delay
            self.config.max_errors_before_stop = max_errors
            
            # Save to file
            save_config(self.config)
            
            self.result = self.config
            self.dialog.destroy()
            
        except ValueError as e:
            messagebox.showerror("Invalid Value", f"Please enter valid numbers for all fields.\n\n{e}")
    
    def _cancel(self):
        """Close dialog without saving."""
        self.result = None
        self.dialog.destroy()
    
    def _reset_defaults(self):
        """Reset all fields to default values."""
        defaults = AppConfig()
        self.printer_var.set(defaults.printer_name)
        self.timeout_var.set(str(defaults.portal_timeout_ms // 1000))
        self.retry_var.set(str(defaults.portal_retry_count))
        self.close_delay_var.set(str(defaults.pdf_close_delay_seconds))
        self.max_errors_var.set(str(defaults.max_errors_before_stop))
    
    def show(self) -> AppConfig | None:
        """Show dialog and return result."""
        self.dialog.wait_window()
        return self.result


class ManifestToolApp:
    """Main application window."""
    
    def __init__(self, root):
        self.root = root
        self.root.title(f"Multi-Carrier Manifest Tool v{__version__}")
        self.root.geometry("750x700")
        self.root.resizable(True, True)
        
        # Set window icon
        self.app_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(self.app_dir, "assets", "icon.ico")
        if os.path.exists(icon_path):
            try:
                self.root.iconbitmap(icon_path)
            except Exception:
                pass  # Icon failed to load, continue without it
        
        # Load config
        self.config = get_config()
        
        # Paths
        self.template_dir = os.path.join(self.app_dir, "templates")
        
        # Variables
        self.carrier_sheet_path = tk.StringVar()
        self.output_dir_path = tk.StringVar()
        self.output_dir_path.set(self.config.default_output_dir)
        self.auto_print_var = tk.BooleanVar(value=True)
        self.auto_upload_var = tk.BooleanVar(value=True)
        
        # Store last generated manifest(s) for printing
        self.last_output_files = []
        self.last_carrier_name = None  # Track which carrier was last processed
        self.last_po_number = None  # Track PO number for Spring upload
        self.last_deutschepost_data = None  # Track Deutsche Post extracted data for portal

        # Batch processing state
        self.batch_files = []  # List of (filepath, carrier_name) tuples
        self.batch_results = []  # List of batch result dicts

        self.create_widgets()
        self.check_templates()
    
    def create_widgets(self):
        """Build the UI."""
        # Main container
        container = ttk.Frame(self.root, padding="10")
        container.grid(row=0, column=0, sticky="nsew")
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)
        container.rowconfigure(1, weight=1)
        
        # Title bar with title and about button
        title_frame = ttk.Frame(container)
        title_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        title_frame.columnconfigure(0, weight=1)
        
        title_label = ttk.Label(
            title_frame, 
            text="Multi-Carrier Manifest Tool", 
            font=('Helvetica', 16, 'bold')
        )
        title_label.grid(row=0, column=0, sticky="w")
        
        # About button (top right, small)
        self.about_button = ttk.Button(
            title_frame,
            text="ℹ",
            width=3,
            command=self.show_about
        )
        self.about_button.grid(row=0, column=1, sticky="e")
        
        # Create notebook (tabbed interface)
        self.notebook = ttk.Notebook(container)
        self.notebook.grid(row=1, column=0, sticky="nsew")
        
        # Tab 1: Manifest Processing
        main_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(main_frame, text="Manifest Processing")
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(5, weight=1)
        
        # File selection frame
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)
        
        # Carrier Sheet
        ttk.Label(file_frame, text="Carrier Sheet:").grid(row=0, column=0, sticky="w", pady=5)
        ttk.Entry(file_frame, textvariable=self.carrier_sheet_path, width=60).grid(
            row=0, column=1, sticky="ew", padx=5
        )
        ttk.Button(file_frame, text="Browse...", command=self.browse_carrier_sheet).grid(
            row=0, column=2
        )
        
        # Output Directory
        ttk.Label(file_frame, text="Output Folder:").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Entry(file_frame, textvariable=self.output_dir_path, width=60).grid(
            row=1, column=1, sticky="ew", padx=5
        )
        ttk.Button(file_frame, text="Browse...", command=self.browse_output_dir).grid(
            row=1, column=2
        )
        
        # Options frame
        options_frame = ttk.LabelFrame(main_frame, text="Options", padding="10")
        options_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 10))
        
        # Auto-print checkbox
        self.auto_print_check = ttk.Checkbutton(
            options_frame,
            text="Auto-print manifest",
            variable=self.auto_print_var
        )
        self.auto_print_check.grid(row=0, column=0, sticky="w")
        
        # Printer info label (shows current printer)
        printer_short = self.config.printer_name.split('\\')[-1] if self.config.printer_name else "default"
        self.printer_info_label = ttk.Label(
            options_frame,
            text=f"Printer: {printer_short}",
            font=('Helvetica', 8),
            foreground='gray'
        )
        self.printer_info_label.grid(row=0, column=1, sticky="w", padx=(10, 0))
        
        # Auto-upload checkbox (Spring, Landmark, Deutsche Post)
        self.auto_upload_check = ttk.Checkbutton(
            options_frame,
            text="Auto-upload to portal (Spring/Landmark/Deutsche Post)",
            variable=self.auto_upload_var
        )
        self.auto_upload_check.grid(row=1, column=0, sticky="w", pady=(5, 0))
        
        # Upload info label
        upload_info = ttk.Label(
            options_frame,
            text="Uses Playwright to upload to carrier portals",
            font=('Helvetica', 8),
            foreground='gray'
        )
        upload_info.grid(row=1, column=1, sticky="w", padx=(20, 0))
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=3, pady=10)
        
        # Process button
        self.process_button = ttk.Button(
            button_frame, 
            text="Process Manifest", 
            command=self.start_processing,
            style='Accent.TButton'
        )
        self.process_button.grid(row=0, column=0, padx=5)
        
        # Print Last button
        self.print_button = ttk.Button(
            button_frame,
            text="Print Last Manifest",
            command=self.print_last_manifest,
            state='disabled'
        )
        self.print_button.grid(row=0, column=1, padx=5)
        
        # Upload button (for manual upload)
        self.upload_button = ttk.Button(
            button_frame,
            text="Upload to Portal",
            command=self.upload_last_manifest,
            state='disabled'
        )
        self.upload_button.grid(row=0, column=2, padx=5)
        
        # Settings button
        self.settings_button = ttk.Button(
            button_frame,
            text="⚙ Settings",
            command=self.open_settings
        )
        self.settings_button.grid(row=0, column=3, padx=5)

        # Batch Process button
        self.batch_button = ttk.Button(
            button_frame,
            text="Batch Process Folder...",
            command=self.start_batch_mode
        )
        self.batch_button.grid(row=0, column=4, padx=5)

        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(0, 10))
        
        # Log frame
        log_frame = ttk.LabelFrame(main_frame, text="Processing Log", padding="5")
        log_frame.grid(row=5, column=0, columnspan=3, sticky="nsew", pady=(0, 10))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(
            log_frame, 
            height=15, 
            width=80,
            font=('Consolas', 9)
        )
        self.log_text.grid(row=0, column=0, sticky="nsew")
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(
            main_frame, 
            textvariable=self.status_var, 
            relief='sunken',
            anchor='w'
        )
        status_bar.grid(row=6, column=0, columnspan=3, sticky="ew")
        
        # Tab 2: Pre-Alerts
        pre_alert_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(pre_alert_frame, text="Pre-Alerts")
        
        # Create pre-alert tab component
        self.pre_alert_tab = PreAlertTab(pre_alert_frame, self.app_dir)
        self.pre_alert_tab.set_log_callback(self.log)
    
    def check_templates(self):
        """Check that template directory exists and has templates."""
        if not os.path.exists(self.template_dir):
            self.log("⚠ Templates folder not found. Creating...")
            os.makedirs(self.template_dir, exist_ok=True)
            self.log(f"  Created: {self.template_dir}")
            self.log("  Please add carrier manifest templates to this folder.")
            return
        
        # Map template filenames to carrier names for display
        template_to_carrier = {
            'Air_Business_Ireland.xlsx': 'Air Business',
            'Asendia_UK_Business_2026_Mail_Manifest.xlsx': 'Asendia 2026',
            'Asendia_UK_Business_Mail_2025.xlsx': 'Asendia 2025',
            'MailOrderTemplate.xlsx': 'Spring',
            'Mail_America_Africa_2025.xlsx': 'Mail Americas',
            'PostNord.xlsx': 'PostNord',
            'UploadCodeList_-_Citipost.xls': 'Landmark Global',
            'United_Business.xlsx': 'United Business ADS',
            'UBL_CP_Pre_Alert_T_D-ETOE.xlsx': 'United Business NZP ETOE',
            'UBL_CP_Pre_Alert_SPL-ETOE.xlsx': 'United Business SPL ETOE',
        }
        
        templates = [f for f in os.listdir(self.template_dir) if f.endswith(('.xlsx', '.xls'))]
        if templates:
            self.log(f"✓ Found {len(templates)} template(s):")
            for t in templates:
                carrier = template_to_carrier.get(t, 'Unknown')
                self.log(f"  - {t} ({carrier})")
        else:
            self.log("⚠ No templates found in templates folder.")
            self.log(f"  Path: {self.template_dir}")
    
    def log(self, message: str):
        """Add message to log."""
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def clear_log(self):
        """Clear the log."""
        self.log_text.delete(1.0, tk.END)
    
    def browse_carrier_sheet(self):
        """Open file dialog for carrier sheet."""
        filepath = filedialog.askopenfilename(
            title="Select Carrier Sheet",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filepath:
            self.carrier_sheet_path.set(filepath)
    
    def browse_output_dir(self):
        """Open folder dialog for output directory."""
        dirpath = filedialog.askdirectory(title="Select Output Folder")
        if dirpath:
            self.output_dir_path.set(dirpath)

    # =========================================================================
    # BATCH PROCESSING METHODS
    # =========================================================================

    def detect_carrier_from_file(self, filepath: str) -> tuple:
        """
        Check if file is a valid carrier sheet by reading B3 cell.
        Returns (is_valid, carrier_name_or_error_message)
        """
        from openpyxl import load_workbook
        from carriers import get_carrier

        if not filepath.lower().endswith(('.xlsx', '.xls')):
            return False, "Not an Excel file"

        try:
            wb = load_workbook(filepath, data_only=True, read_only=True)
            ws = wb.active
            carrier_name = str(ws['B3'].value or "").strip()
            wb.close()

            if not carrier_name:
                return False, "No carrier name in B3"

            # Validate carrier is known
            get_carrier(carrier_name)
            return True, carrier_name
        except ValueError as e:
            return False, str(e)
        except Exception as e:
            return False, f"Error reading file: {str(e)}"

    def start_batch_mode(self):
        """Open folder dialog and scan for carrier sheets."""
        # Validate output directory first
        output_dir = self.output_dir_path.get()
        if not output_dir or not os.path.isdir(output_dir):
            messagebox.showerror("Error", "Please select a valid output folder first.")
            return

        # Open folder dialog
        folder = filedialog.askdirectory(title="Select Folder with Carrier Sheets")
        if not folder:
            return

        # Scan folder for Excel files
        self.log(f"\nScanning folder: {folder}")
        excel_files = []
        for f in sorted(os.listdir(folder)):
            if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$'):
                excel_files.append(os.path.join(folder, f))

        if not excel_files:
            messagebox.showinfo("No Files", "No Excel files found in folder.")
            return

        # Detect carriers for each file
        self.batch_files = []
        skipped = []
        for filepath in excel_files:
            is_valid, result = self.detect_carrier_from_file(filepath)
            if is_valid:
                self.batch_files.append((filepath, result))
                self.log(f"  + {os.path.basename(filepath)} - {result}")
            else:
                skipped.append((os.path.basename(filepath), result))
                self.log(f"  - {os.path.basename(filepath)} - {result}")

        if not self.batch_files:
            messagebox.showinfo("No Valid Files", "No valid carrier sheets found.")
            return

        # Check for duplicate carriers (same carrier type appearing multiple times)
        carrier_counts = {}
        for filepath, carrier_name in self.batch_files:
            # Normalize carrier name for comparison (the resolved carrier type)
            try:
                from carriers import get_carrier
                carrier_instance = get_carrier(carrier_name)
                carrier_type = type(carrier_instance).__name__
            except ValueError:
                carrier_type = carrier_name

            if carrier_type not in carrier_counts:
                carrier_counts[carrier_type] = []
            carrier_counts[carrier_type].append(os.path.basename(filepath))

        # Find duplicates
        duplicates = {k: v for k, v in carrier_counts.items() if len(v) > 1}
        if duplicates:
            dup_msg = "Duplicate carrier sheets detected!\n\n"
            dup_msg += "Each carrier type should only appear once per batch.\n\n"
            for carrier_type, files in duplicates.items():
                dup_msg += f"{carrier_type}:\n"
                for f in files:
                    dup_msg += f"  - {f}\n"
            dup_msg += "\nPlease remove duplicate files and try again."
            messagebox.showerror("Duplicate Carriers", dup_msg)
            return

        # Confirm with user
        msg = f"Found {len(self.batch_files)} valid carrier sheet(s):\n\n"
        for fp, carrier in self.batch_files[:10]:  # Show first 10
            msg += f"  {os.path.basename(fp)} ({carrier})\n"
        if len(self.batch_files) > 10:
            msg += f"... and {len(self.batch_files) - 10} more\n"
        if skipped:
            msg += f"\n{len(skipped)} file(s) skipped (invalid/unknown carrier)"
        msg += "\n\nProceed with batch processing?"

        if messagebox.askyesno("Confirm Batch Processing", msg):
            self.start_batch_processing()

    def start_batch_processing(self):
        """Start processing all detected batch files."""
        if not self.batch_files:
            return

        # Disable buttons during processing
        self.process_button.config(state='disabled')
        self.batch_button.config(state='disabled')
        self.print_button.config(state='disabled')
        self.upload_button.config(state='disabled')

        # Start progress bar
        self.progress.start()
        self.batch_results = []

        # Run in background thread
        thread = threading.Thread(
            target=self.run_batch_processing,
            daemon=True
        )
        thread.start()

    def run_batch_processing(self):
        """Background thread for batch processing."""
        output_dir = self.output_dir_path.get()
        total = len(self.batch_files)

        for index, (filepath, carrier_name) in enumerate(self.batch_files):
            filename = os.path.basename(filepath)

            # Update status (capture index and filename in closure)
            self.root.after(0, lambda i=index, f=filename, t=total:
                self.status_var.set(f"Batch: Processing {i+1}/{t} - {f}"))

            self.root.after(0, self.log, f"\n{'='*50}")
            self.root.after(0, self.log, f"BATCH FILE {index+1}/{total}: {filename}")
            self.root.after(0, self.log, f"Carrier: {carrier_name}")
            self.root.after(0, self.log, f"{'='*50}")

            try:
                # Create engine and process
                engine = ManifestEngine(self.template_dir, output_dir)
                engine.set_log_callback(lambda msg: self.root.after(0, self.log, msg))

                results = engine.process_sheet(filepath, max_errors=self.config.max_errors_before_stop)

                if results and results[0].success:
                    # Handle auto-print (skip for Spring/Landmark - portal handles it)
                    is_spring = 'spring' in carrier_name.lower()
                    is_landmark = 'landmark' in carrier_name.lower()
                    is_deutschepost = 'deutsche' in carrier_name.lower()

                    if self.auto_print_var.get() and not is_spring and not is_landmark:
                        self._do_print([results[0].output_file])

                    # Handle auto-upload
                    if self.auto_upload_var.get():
                        self._handle_batch_upload(results, carrier_name, output_dir, is_spring, is_landmark, is_deutschepost)

                    self.batch_results.append({
                        'file': filename,
                        'carrier': carrier_name,
                        'success': True,
                        'output': results[0].output_file,
                        'po_number': results[0].po_number
                    })
                else:
                    self.batch_results.append({
                        'file': filename,
                        'carrier': carrier_name,
                        'success': False,
                        'error': 'Processing failed'
                    })

            except Exception as e:
                self.root.after(0, self.log, f"ERROR: {str(e)}")
                self.batch_results.append({
                    'file': filename,
                    'carrier': carrier_name,
                    'success': False,
                    'error': str(e)
                })

        # Complete
        self.root.after(0, self.on_batch_complete)

    def _handle_batch_upload(self, results, carrier_name, output_dir, is_spring, is_landmark, is_deutschepost):
        """Handle auto-upload for a single batch file."""
        if is_spring:
            self._upload_spring_blocking(results[0].output_file, results[0].po_number, output_dir)
        elif is_landmark:
            files = [results[0].output_file]
            if results[0].additional_files:
                files.extend(results[0].additional_files)
            self._upload_landmark_blocking(files, results[0].po_number, output_dir)
        elif is_deutschepost and results[0].deutschepost_data:
            self._upload_deutschepost_blocking(results[0], output_dir)

    def _upload_spring_blocking(self, output_file, po_number, output_dir):
        """Synchronous Spring upload for batch mode."""
        from carriers.spring_portal import run_spring_upload_robust
        try:
            success, msg, pdf_downloaded = run_spring_upload_robust(
                output_file, po_number, output_dir,
                self.auto_print_var.get(),
                lambda m: self.root.after(0, self.log, m)
            )
            self.root.after(0, self.log, f"Spring upload: {'OK' if success else 'FAILED'}")

            # Delete upload file if PDF was successfully downloaded
            if success and pdf_downloaded:
                try:
                    os.remove(output_file)
                    self.root.after(0, self.log, f"  ✓ Cleaned up upload file: {os.path.basename(output_file)}")
                except Exception as e:
                    self.root.after(0, self.log, f"  ⚠ Could not delete upload file {os.path.basename(output_file)}: {e}")
            elif success:
                self.root.after(0, self.log, "  ⚠ Upload file retained (PDF download incomplete)")
        except Exception as e:
            self.root.after(0, self.log, f"Spring upload error: {e}")

    def _upload_landmark_blocking(self, files, po_number, output_dir):
        """Synchronous Landmark upload for batch mode."""
        # Run async function in new event loop
        loop = asyncio.new_event_loop()
        try:
            success, message, pdf_downloaded = loop.run_until_complete(
                upload_to_landmark_portal(files, po_number, output_dir,
                                          self.auto_print_var.get(),
                                          lambda m: self.root.after(0, self.log, m))
            )
            self.root.after(0, self.log, f"Landmark upload: {'OK' if success else 'FAILED'}")

            # Delete upload CSV files if PDFs were successfully downloaded
            if success and pdf_downloaded:
                for filepath in files:
                    try:
                        os.remove(filepath)
                        self.root.after(0, self.log, f"  ✓ Cleaned up upload file: {os.path.basename(filepath)}")
                    except Exception as e:
                        self.root.after(0, self.log, f"  ⚠ Could not delete upload file {os.path.basename(filepath)}: {e}")
            elif success:
                self.root.after(0, self.log, "  ⚠ Upload files retained (PDF download incomplete)")
        except Exception as e:
            self.root.after(0, self.log, f"Landmark upload error: {e}")
        finally:
            loop.close()

    def _upload_deutschepost_blocking(self, result, output_dir):
        """Synchronous Deutsche Post upload for batch mode."""
        from carriers.deutschepost_portal import run_deutschepost_upload
        from carriers.deutschepost import DeutschePostCarrier
        try:
            data = result.deutschepost_data
            carrier = DeutschePostCarrier()
            item_format = carrier.get_item_format(data.formats)

            success, msg = run_deutschepost_upload(
                po_number=data.po_number,
                total_weight=data.total_weight,
                item_format=item_format,
                output_dir=output_dir,
                auto_print=self.auto_print_var.get(),
                log_callback=lambda m: self.root.after(0, self.log, m)
            )
            self.root.after(0, self.log, f"Deutsche Post upload: {'OK' if success else 'FAILED'}")
        except Exception as e:
            self.root.after(0, self.log, f"Deutsche Post upload error: {e}")

    def on_batch_complete(self):
        """Handle batch processing completion."""
        self.progress.stop()
        self.process_button.config(state='normal')
        self.batch_button.config(state='normal')

        # Calculate summary
        successful = sum(1 for r in self.batch_results if r['success'])
        failed = len(self.batch_results) - successful

        # Log summary
        self.log(f"\n{'='*50}")
        self.log("BATCH PROCESSING COMPLETE")
        self.log(f"{'='*50}")
        self.log(f"Total: {len(self.batch_results)}")
        self.log(f"Successful: {successful}")
        self.log(f"Failed: {failed}")

        for r in self.batch_results:
            status = "OK" if r['success'] else "FAILED"
            self.log(f"  [{status}] {r['file']} ({r['carrier']})")
            if not r['success']:
                self.log(f"      Error: {r.get('error', 'Unknown')}")

        self.status_var.set(f"Batch complete: {successful} succeeded, {failed} failed")

        # Add successful manifests to pre-alert queue
        for r in self.batch_results:
            if r['success'] and r.get('output'):
                self.pre_alert_tab.add_manifest(
                    carrier_name=r['carrier'],
                    po_number=r.get('po_number', ''),
                    manifest_path=r['output']
                )

        # Show summary dialog
        messagebox.showinfo(
            "Batch Processing Complete",
            f"Processed {len(self.batch_results)} file(s)\n\n"
            f"Successful: {successful}\n"
            f"Failed: {failed}"
        )

        # Clear batch state
        self.batch_files = []

    # =========================================================================
    # END BATCH PROCESSING METHODS
    # =========================================================================

    def open_settings(self):
        """Open settings dialog."""
        dialog = SettingsDialog(self.root, self.config)
        result = dialog.show()
        
        if result:
            # Config was saved - update UI to reflect changes
            self.config = result
            
            # Update printer label
            printer_short = self.config.printer_name.split('\\')[-1] if self.config.printer_name else "default"
            self.printer_info_label.config(text=f"Printer: {printer_short}")
            
            self.log(f"Settings saved. Printer: {printer_short}")
    
    def show_about(self):
        """Show about dialog with version and changelog info."""
        # Read changelog
        changelog_path = os.path.join(self.app_dir, "CHANGELOG.md")
        changelog_content = ""
        
        if os.path.exists(changelog_path):
            try:
                with open(changelog_path, 'r', encoding='utf-8') as f:
                    changelog_content = f.read()
            except Exception:
                changelog_content = "Could not load changelog."
        
        # Create about dialog
        about_dialog = tk.Toplevel(self.root)
        about_dialog.title("About")
        about_dialog.geometry("550x550")
        about_dialog.resizable(False, False)
        about_dialog.transient(self.root)
        about_dialog.grab_set()
        
        # Center on parent
        about_dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (550 // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (450 // 2)
        about_dialog.geometry(f"+{x}+{y}")
        
        # Main frame
        frame = ttk.Frame(about_dialog, padding="10")
        frame.pack(fill='both', expand=True)
        
        # Create notebook (tabs)
        notebook = ttk.Notebook(frame)
        notebook.pack(fill='both', expand=True)
        
        # === About Tab ===
        about_tab = ttk.Frame(notebook, padding="10")
        notebook.add(about_tab, text="About")
        
        about_text = f"""Multi-Carrier Manifest Tool

Version: {__version__}
Author: {__author__}

Automates population of carrier manifests from internal
carrier sheets with integrated workflow automation.

Features:
• Auto-detection of carrier from carrier sheet
• Batch processing of multiple carrier sheets
• Portal automation (Spring, Landmark, Deutsche Post)
• Pre-alert email automation via Outlook
• Manifest queue with per-carrier email configuration
• Auto-print to configured printer
• Configurable settings

© 2026 Citipost Ltd"""
        
        about_label = ttk.Label(about_tab, text=about_text, justify='left')
        about_label.pack(anchor='w')
        
        # === Supported Carriers Tab ===
        carriers_tab = ttk.Frame(notebook, padding="10")
        notebook.add(carriers_tab, text="Supported Carriers")

        carriers_content = """Supported Carriers:

• Asendia UK Business Mail (2025)
   Template: Asendia_UK_Business_Mail_2025.xlsx

• Asendia UK Business Mail (2026)
   Template: Asendia_UK_Business_2026_Mail_Manifest.xlsx

• PostNord Business Mail
   Template: PostNord.xlsx

• Spring Global Delivery Solutions
   Template: MailOrderTemplate.xlsx
   Portal: Automatic upload and manifest download

• Landmark Global
   Template: UploadCodeList_-_Citipost.xls
   Portal: Automatic CSV upload (Economy & Priority)

• Deutsche Post
   Portal: Automatic registration and manifest download

• Air Business (Ireland / An Post)
   Template: Air_Business_Ireland.xlsx

• Mail Americas/Africa
   Template: Mail_America_Africa_2025.xlsx

• United Business ADS
   Template: United_Business.xlsx

• United Business NZP ETOE
   Template: UBL_CP_Pre_Alert_T_D-ETOE.xlsx

• United Business SPL ETOE
   Template: UBL_CP_Pre_Alert_SPL-ETOE.xlsx
"""

        carriers_text = scrolledtext.ScrolledText(carriers_tab, wrap='word', font=('Consolas', 9))
        carriers_text.pack(fill='both', expand=True)
        carriers_text.insert('1.0', carriers_content)
        carriers_text.config(state='disabled')
        
        # === Changelog Tab ===
        changelog_tab = ttk.Frame(notebook, padding="10")
        notebook.add(changelog_tab, text="Changelog")
        
        changelog_text = scrolledtext.ScrolledText(changelog_tab, wrap='word', font=('Consolas', 9))
        changelog_text.pack(fill='both', expand=True)
        changelog_text.insert('1.0', changelog_content)
        changelog_text.config(state='disabled')
        
        # Close button
        ttk.Button(frame, text="Close", command=about_dialog.destroy).pack(pady=(10, 0))
    
    def start_processing(self):
        """Start processing in background thread."""
        carrier_sheet = self.carrier_sheet_path.get()
        output_dir = self.output_dir_path.get()
        
        # Validation
        if not carrier_sheet:
            messagebox.showerror("Error", "Please select a carrier sheet.")
            return
        
        if not os.path.exists(carrier_sheet):
            messagebox.showerror("Error", f"Carrier sheet not found:\n{carrier_sheet}")
            return
        
        if not output_dir:
            messagebox.showerror("Error", "Please select an output folder.")
            return
        
        if not os.path.exists(output_dir):
            messagebox.showerror("Error", f"Output folder not found:\n{output_dir}")
            return
        
        # Disable buttons, start progress
        self.process_button.config(state='disabled')
        self.print_button.config(state='disabled')
        self.upload_button.config(state='disabled')
        self.progress.start()
        self.status_var.set("Processing...")
        self.clear_log()
        
        # Clear previous output files
        self.last_output_files = []
        self.last_carrier_name = None
        self.last_po_number = None
        self.last_deutschepost_data = None
        
        # Run in background thread
        thread = threading.Thread(
            target=self.run_processing,
            args=(carrier_sheet, output_dir),
            daemon=True
        )
        thread.start()
    
    def run_processing(self, carrier_sheet: str, output_dir: str):
        """Background processing task."""
        try:
            engine = ManifestEngine(self.template_dir, output_dir)
            engine.set_log_callback(lambda msg: self.root.after(0, self.log, msg))
            
            results = engine.process_sheet(carrier_sheet, max_errors=self.config.max_errors_before_stop)
            
            # Report results
            self.root.after(0, self.on_processing_complete, results)
            
        except Exception as e:
            self.root.after(0, self.on_processing_error, str(e))
    
    def on_processing_complete(self, results):
        """Handle processing completion."""
        self.progress.stop()
        self.process_button.config(state='normal')
        
        # Collect output files, carrier name, and PO number
        # Include both primary output_file and any additional_files (e.g., Landmark Economy + Priority)
        self.last_output_files = []
        for r in results:
            if r.success and r.output_file:
                self.last_output_files.append(r.output_file)
                self.log(f"Collected primary output: {os.path.basename(r.output_file)}")
            # Also add any additional files (e.g., Landmark Priority CSV)
            if hasattr(r, 'additional_files') and r.additional_files:
                self.last_output_files.extend(r.additional_files)
                self.log(f"Collected additional files: {[os.path.basename(f) for f in r.additional_files]}")
        
        self.log(f"Total files collected for upload/print: {len(self.last_output_files)}")
        for f in self.last_output_files:
            self.log(f"  - {os.path.basename(f)}")
        
        if results:
            self.last_carrier_name = results[0].carrier_name
            self.last_po_number = results[0].po_number
            # Store Deutsche Post data if available
            if hasattr(results[0], 'deutschepost_data') and results[0].deutschepost_data:
                self.last_deutschepost_data = results[0].deutschepost_data
        
        # Enable print button if we have output files
        if self.last_output_files:
            self.print_button.config(state='normal')
            # Enable upload button for Spring, Landmark, or Deutsche Post
            if self.last_carrier_name and ('spring' in self.last_carrier_name.lower() or 'landmark' in self.last_carrier_name.lower() or 'deutsche' in self.last_carrier_name.lower()):
                self.upload_button.config(state='normal')
        
        # Summarise results
        self.log("\n" + "="*50)
        self.log("PROCESSING COMPLETE")
        self.log("="*50)
        
        total_processed = 0
        total_failed = 0
        all_errors = []
        
        for result in results:
            self.log(f"\n{result.carrier_name}:")
            self.log(f"  Records processed: {result.records_processed}")
            self.log(f"  Records failed: {result.records_failed}")
            
            if result.output_file:
                self.log(f"  Output: {os.path.basename(result.output_file)}")
            
            total_processed += result.records_processed
            total_failed += result.records_failed
            all_errors.extend(result.errors)
        
        self.log(f"\nTOTAL: {total_processed} processed, {total_failed} failed")
        
        # Auto-print if enabled and successful
        # Skip for Spring/Landmark - portal handles printing
        # For Deutsche Post, print the carrier sheet now, portal will print the manifest PDF later
        is_spring = self.last_carrier_name and 'spring' in self.last_carrier_name.lower()
        is_landmark = self.last_carrier_name and 'landmark' in self.last_carrier_name.lower()
        is_deutschepost = self.last_carrier_name and 'deutsche' in self.last_carrier_name.lower()
        if self.auto_print_var.get() and self.last_output_files and not is_spring and not is_landmark:
            self.log("\n" + "-"*50)
            self.log("AUTO-PRINT ENABLED")
            self._do_print(self.last_output_files)
        
        # Auto-upload if enabled, successful, and Spring or Landmark carrier
        if (self.auto_upload_var.get() and 
            self.last_output_files and 
            is_spring):
            self.log("\n" + "-"*50)
            self.log("AUTO-UPLOAD TO SPRING PORTAL")
            self._do_upload_spring(self.last_output_files[0], self.last_po_number, self.output_dir_path.get(), self.auto_print_var.get())
        
        if (self.auto_upload_var.get() and 
            self.last_output_files and 
            is_landmark):
            self.log("\n" + "-"*50)
            self.log("AUTO-UPLOAD TO LANDMARK PORTAL")
            self._do_upload_landmark(self.last_output_files, self.last_po_number, self.output_dir_path.get(), self.auto_print_var.get())
        
        if (self.auto_upload_var.get() and 
            self.last_output_files and 
            is_deutschepost):
            self.log("\n" + "-"*50)
            self.log("AUTO-UPLOAD TO DEUTSCHE POST PORTAL")
            # For Deutsche Post, we need to extract weight and format from the carrier sheet
            self._do_upload_deutschepost(self.last_po_number, self.output_dir_path.get(), self.auto_print_var.get())
        
        if all_errors:
            self.status_var.set(f"Completed with {len(all_errors)} error(s)")
            
            # Show error summary
            error_summary = "\n".join(f"• {e}" for e in all_errors[:10])
            if len(all_errors) > 10:
                error_summary += f"\n... and {len(all_errors) - 10} more"
            
            messagebox.showwarning(
                "Completed with Errors",
                f"Processing completed with {len(all_errors)} error(s):\n\n{error_summary}"
            )
        else:
            self.status_var.set("Completed successfully")
            msg = (
                f"Manifest populated successfully!\n\n"
                f"Records processed: {total_processed}\n"
                f"Output folder: {os.path.dirname(results[0].output_file) if results and results[0].output_file else 'N/A'}"
            )
            if self.auto_print_var.get() and not is_spring and not is_landmark:
                printer_short = self.config.printer_name.split('\\')[-1] if self.config.printer_name else "default"
                msg += f"\n\nSent to printer: {printer_short}"
            if (self.auto_upload_var.get() and is_spring):
                msg += "\n\nUpload to Spring portal initiated."
            if (self.auto_upload_var.get() and is_landmark):
                msg += "\n\nUpload to Landmark portal initiated."
            if (self.auto_upload_var.get() and is_deutschepost):
                msg += "\n\nUpload to Deutsche Post portal initiated."
            messagebox.showinfo("Success", msg)
        
        # Add to pre-alerts tab if it's a pre-alert carrier
        if results and results[0].success and results[0].output_file:
            self.pre_alert_tab.add_manifest(
                carrier_name=results[0].carrier_name,
                po_number=results[0].po_number,
                manifest_path=results[0].output_file
            )
    
    def on_processing_error(self, error_msg: str):
        """Handle processing error."""
        self.progress.stop()
        self.process_button.config(state='normal')
        self.status_var.set("Error occurred")
        
        self.log(f"\n✗ ERROR: {error_msg}")
        messagebox.showerror("Error", f"Processing failed:\n\n{error_msg}")
    
    def print_last_manifest(self):
        """Print the last generated manifest(s)."""
        if not self.last_output_files:
            messagebox.showwarning("No Manifest", "No manifest available to print.")
            return
        
        # Confirm print
        file_list = "\n".join(f"• {os.path.basename(f)}" for f in self.last_output_files)
        printer_short = self.config.printer_name.split("\\")[-1] if self.config.printer_name else "default"
        
        if not messagebox.askyesno(
            "Confirm Print",
            f"Print the following manifest(s) to {printer_short}?\n\n{file_list}"
        ):
            return
        
        self._do_print(self.last_output_files)
    
    def _do_print(self, files: list):
        """Execute print for given files."""
        self.log("-"*50)
        
        for filepath in files:
            filename = os.path.basename(filepath)
            self.log(f"Printing: {filename}...")
            
            success, message = print_excel_workbook(filepath, self.config.printer_name)
            
            if success:
                self.log(f"  ✓ {message}")
            else:
                self.log(f"  ✗ {message}")
                messagebox.showerror(
                    "Print Failed",
                    f"Failed to print {filename}:\n\n{message}"
                )
    
    def upload_last_manifest(self):
        """Upload the last generated manifest to carrier portal."""
        if not self.last_output_files:
            messagebox.showwarning("No Manifest", "No manifest available to upload.")
            return
        
        is_spring = self.last_carrier_name and 'spring' in self.last_carrier_name.lower()
        is_landmark = self.last_carrier_name and 'landmark' in self.last_carrier_name.lower()
        is_deutschepost = self.last_carrier_name and 'deutsche' in self.last_carrier_name.lower()
        
        if not is_spring and not is_landmark and not is_deutschepost:
            messagebox.showwarning(
                "Upload Not Available",
                "Upload to portal is only available for Spring, Landmark, and Deutsche Post manifests."
            )
            return
        
        # Confirm upload
        file_list = "\n".join(f"• {os.path.basename(f)}" for f in self.last_output_files)
        if is_spring:
            portal_name = "MySpring"
        elif is_landmark:
            portal_name = "Landmark"
        else:
            portal_name = "Deutsche Post"
        
        if not messagebox.askyesno(
            "Confirm Upload",
            f"Upload the following to {portal_name} portal?\n\n{file_list}"
        ):
            return
        
        if is_spring:
            self._do_upload_spring(self.last_output_files[0], self.last_po_number, self.output_dir_path.get(), self.auto_print_var.get())
        elif is_landmark:
            self._do_upload_landmark(self.last_output_files, self.last_po_number, self.output_dir_path.get(), self.auto_print_var.get())
        else:
            self._do_upload_deutschepost(self.last_po_number, self.output_dir_path.get(), self.auto_print_var.get())
    
    def _do_upload_spring(self, filepath: str, po_number: str = "", output_dir: str = "", auto_print: bool = True):
        """Execute Spring upload in background thread."""
        self.progress.start()
        self.status_var.set("Uploading to Spring portal...")
        self.upload_button.config(state='disabled')
        
        # Run upload in background thread
        thread = threading.Thread(
            target=self._upload_spring_thread,
            args=(filepath, po_number, output_dir, auto_print),
            daemon=True
        )
        thread.start()
    
    def _upload_spring_thread(self, filepath: str, po_number: str = "", output_dir: str = "", auto_print: bool = True):
        """Background Spring upload task."""
        filename = os.path.basename(filepath)
        
        def log_msg(msg):
            self.root.after(0, self.log, msg)
        
        log_msg(f"Uploading: {filename}")
        if po_number:
            log_msg(f"  PO Number: {po_number}")
        if output_dir:
            log_msg(f"  Output folder: {output_dir}")
        
        success, message, pdf_downloaded = run_spring_upload(filepath, po_number, output_dir, auto_print, log_msg)
        
        # Delete upload file if PDF was successfully downloaded and printed
        if success and pdf_downloaded:
            try:
                os.remove(filepath)
                log_msg(f"  ✓ Cleaned up upload file: {filename}")
            except Exception as e:
                log_msg(f"  ⚠ Could not delete upload file: {e}")
        elif success:
            log_msg(f"  ⚠ Upload file retained for debugging: {filename}")
        
        # Update UI on main thread
        self.root.after(0, self._on_upload_complete, success, message)
    
    def _do_upload_landmark(self, filepaths: list, po_number: str = "", output_dir: str = "", auto_print: bool = True):
        """Execute Landmark upload in background thread."""
        self.progress.start()
        self.status_var.set("Uploading to Landmark portal...")
        self.upload_button.config(state='disabled')
        
        # Run upload in background thread
        thread = threading.Thread(
            target=self._upload_landmark_thread,
            args=(filepaths, po_number, output_dir, auto_print),
            daemon=True
        )
        thread.start()
    
    def _upload_landmark_thread(self, filepaths: list, po_number: str = "", output_dir: str = "", auto_print: bool = True):
        """Background Landmark upload task."""
        def log_msg(msg):
            self.root.after(0, self.log, msg)
        
        log_msg(f"Uploading {len(filepaths)} file(s) to Landmark portal...")
        for f in filepaths:
            log_msg(f"  - {os.path.basename(f)}")
        if po_number:
            log_msg(f"  PO Number: {po_number}")
        if output_dir:
            log_msg(f"  Output folder: {output_dir}")
        
        success, message, pdf_downloaded = run_landmark_upload(filepaths, po_number, output_dir, auto_print, log_msg)
        
        # Delete upload files (CSV) if PDFs were successfully downloaded and printed
        if success and pdf_downloaded:
            for filepath in filepaths:
                try:
                    os.remove(filepath)
                    log_msg(f"  ✓ Cleaned up upload file: {os.path.basename(filepath)}")
                except Exception as e:
                    log_msg(f"  ⚠ Could not delete upload file {os.path.basename(filepath)}: {e}")
        elif success:
            log_msg("  ⚠ Upload files retained for debugging")
        
        # Update UI on main thread
        self.root.after(0, self._on_upload_complete, success, message)
    
    def _do_upload_deutschepost(self, po_number: str = "", output_dir: str = "", auto_print: bool = True):
        """Execute Deutsche Post portal registration in background thread."""
        # Check if we have the required data
        if not self.last_deutschepost_data:
            messagebox.showerror(
                "Missing Data",
                "Deutsche Post data not available. Please re-process the carrier sheet."
            )
            return
        
        self.progress.start()
        self.status_var.set("Registering on Deutsche Post portal...")
        self.upload_button.config(state='disabled')
        
        # Run upload in background thread
        thread = threading.Thread(
            target=self._upload_deutschepost_thread,
            args=(po_number, output_dir, auto_print),
            daemon=True
        )
        thread.start()
    
    def _upload_deutschepost_thread(self, po_number: str = "", output_dir: str = "", auto_print: bool = True):
        """Background Deutsche Post portal task."""
        from carriers.deutschepost_portal import run_deutschepost_upload
        from carriers.deutschepost import DeutschePostCarrier
        
        def log_msg(msg):
            self.root.after(0, self.log, msg)
        
        # Get data from last processing
        data = self.last_deutschepost_data
        carrier = DeutschePostCarrier()
        item_format = carrier.get_item_format(data.formats)
        
        log_msg("Registering on Deutsche Post portal...")
        log_msg(f"  PO Number: {data.po_number}")
        log_msg(f"  Total Weight: {data.total_weight} kg")
        log_msg(f"  Item Format: {item_format}")
        log_msg(f"  Output folder: {output_dir}")
        
        success, message = run_deutschepost_upload(
            po_number=data.po_number,
            total_weight=data.total_weight,
            item_format=item_format,
            output_dir=output_dir,
            auto_print=auto_print,
            log_callback=log_msg
        )
        
        # Update UI on main thread
        self.root.after(0, self._on_upload_complete, success, message)
    
    def _on_upload_complete(self, success: bool, message: str):
        """Handle upload completion."""
        self.progress.stop()
        self.upload_button.config(state='normal')
        
        if success:
            self.log(f"  ✓ {message}")
            self.status_var.set("Upload completed")
            messagebox.showinfo("Upload Complete", message)
        else:
            self.log(f"  ✗ {message}")
            self.status_var.set("Upload failed")
            messagebox.showerror("Upload Failed", message)


def main():
    root = tk.Tk()
    
    # Try to set a nicer theme
    try:
        style = ttk.Style()
        if 'clam' in style.theme_names():
            style.theme_use('clam')
    except Exception:
        pass
    
    ManifestToolApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
